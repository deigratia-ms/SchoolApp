from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, F, Avg, Max, Min
from django.http import JsonResponse
import json

from .models import Assignment, Question, Choice, StudentSubmission, StudentAnswer, Grade, ReportCard, QuizAttempt, AssessmentWeightConfiguration
from users.models import Student, Parent, Teacher
from courses.models import ClassSubject, ClassRoom

# Helper function for teacher check
def is_teacher(user):
    return user.is_authenticated and user.role == 'TEACHER'

# Helper function for admin check
def is_admin(user):
    return user.is_authenticated and user.role == 'ADMIN'

# Helper function for student check
def is_student(user):
    return user.is_authenticated and user.role == 'STUDENT'

# Helper function for parent check
def is_parent(user):
    return user.is_authenticated and user.role == 'PARENT'

# Assignment views
@login_required
def assignment_list(request):
    """
    Display a list of assignments based on user role:
    - Teachers see assignments they created
    - Students see assignments for their classes
    - Parents see their children's assignments
    - Admins see all assignments
    """
    user = request.user
    assignments = []
    assignment_type_filter = request.GET.get('type', 'all')

    if is_teacher(user):
        # Teachers see assignments they created or are assigned to
        try:
            teacher = Teacher.objects.get(user=user)
            # Check if this is a subject teacher (not a class teacher for any class)
            is_subject_teacher = not ClassRoom.objects.filter(class_teacher=teacher).exists()
            if is_subject_teacher:
                # Filter assignments to only include those for subjects taught by the teacher
                assignments = Assignment.objects.filter(class_subject__teacher=teacher)
            else:
                # Class teachers see all assignments for their classes
                class_teacher_of = ClassRoom.objects.filter(class_teacher=teacher)
                assignments = Assignment.objects.filter(class_subject__classroom__in=class_teacher_of)
        except Teacher.DoesNotExist:
            assignments = []  # Handle case where teacher profile doesn't exist

    elif is_student(user):
        # Students see assignments for their enrolled classes with submission status
        try:
            student = Student.objects.get(user=user)
            enrolled_subjects = ClassSubject.objects.filter(students=student)
            assignments = Assignment.objects.filter(class_subject__in=enrolled_subjects)

            # Get submissions for this student to show completion status
            student_submissions = StudentSubmission.objects.filter(student=student)
            submission_map = {sub.assignment_id: sub for sub in student_submissions}

            # Add submission status to each assignment
            assignments_with_status = []
            for assignment in assignments:
                assignment.submission = submission_map.get(assignment.id)
                assignments_with_status.append(assignment)
            assignments = assignments_with_status

        except Student.DoesNotExist:
            pass

    elif is_parent(user):
        # Parents see their children's assignments
        try:
            parent = Parent.objects.get(user=user)
            children = parent.children.all()

            # Get assignments for all children
            assignments = []
            for child in children:
                enrolled_subjects = ClassSubject.objects.filter(students=child)
                child_assignments = Assignment.objects.filter(class_subject__in=enrolled_subjects)
                assignments.extend(list(child_assignments))

            # Remove duplicates
            assignments = list({assignment.id: assignment for assignment in assignments}.values())
        except Parent.DoesNotExist:
            pass

    elif is_admin(user):
        # Admins see all assignments
        assignments = Assignment.objects.all()

    # Apply assignment type filter if specified
    if assignment_type_filter != 'all' and assignments:
        assignments = [a for a in assignments if a.assignment_type == assignment_type_filter]

    # Order assignments by due date
    assignments = sorted(assignments, key=lambda a: a.due_date)

    context = {
        'assignments': assignments,
        'user': user,
        'filter_type': assignment_type_filter
    }

    return render(request, 'assignments/assignment_list.html', context)

@login_required
@user_passes_test(lambda u: is_teacher(u) or is_admin(u))
def create_assignment(request):
    """
    Create a new assignment. Only class teachers and admins can create assignments.
    """
    # Get the classes that the teacher is teaching
    class_subjects = []

    if is_teacher(request.user):
        try:
            teacher = Teacher.objects.get(user=request.user)

            # Get all class subjects this teacher is assigned to teach
            # This includes subjects where they are either:
            # 1. The class teacher of the classroom, or
            # 2. The subject teacher assigned to the class subject
            class_subjects = ClassSubject.objects.filter(
                Q(classroom__class_teacher=teacher) |
                Q(teacher=teacher)
            ).distinct()

            # If no class subjects are found, show appropriate message
            if not class_subjects.exists():
                messages.error(request, "You are not assigned to teach any subjects.")
                return redirect('assignments:assignment_list')
        except Teacher.DoesNotExist:
            messages.error(request, "Teacher profile not found.")
            return redirect('dashboard:index')
    elif is_admin(request.user):
        class_subjects = ClassSubject.objects.all()

    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        class_subject_id = request.POST.get('class_subject')
        assignment_type = request.POST.get('assignment_type')
        max_score = request.POST.get('max_score')
        due_date = request.POST.get('due_date')
        due_time = request.POST.get('due_time')
        is_active = 'is_active' in request.POST

        # Combine date and time
        from datetime import datetime
        due_datetime_str = f"{due_date} {due_time}"
        due_datetime = datetime.strptime(due_datetime_str, "%Y-%m-%d %H:%M")

        # Validate inputs
        if not (title and description and class_subject_id and assignment_type and max_score and due_date and due_time):
            messages.error(request, "Please fill in all required fields.")
            return render(request, 'assignments/create_assignment.html', {
                'class_subjects': class_subjects,
            })

        try:
            class_subject = ClassSubject.objects.get(id=class_subject_id)

            # Get time limit for quizzes
            time_limit = None
            if assignment_type == 'QUIZ':
                time_limit = request.POST.get('time_limit')
                if time_limit and time_limit.strip():
                    time_limit = int(time_limit)
                else:
                    time_limit = None

            # Create assignment
            assignment = Assignment.objects.create(
                title=title,
                description=description,
                class_subject=class_subject,
                assignment_type=assignment_type,
                max_score=max_score,
                due_date=due_datetime,
                is_active=is_active,
                created_by=request.user,
                time_limit=time_limit
            )

            # Handle file attachment if provided
            if 'file_attachment' in request.FILES:
                assignment.file_attachment = request.FILES['file_attachment']
                assignment.save()

            messages.success(request, f"Assignment '{title}' created successfully.")

            # Create notifications for students and parents
            try:
                from communications.utils import create_assignment_notifications
                create_assignment_notifications(assignment)
            except Exception as e:
                print(f"Error creating notifications: {e}")

            # Redirect based on assignment type
            if assignment_type == 'QUIZ':
                return redirect('assignments:create_question', assignment_id=assignment.id)
            else:
                return redirect('assignments:assignment_detail', assignment_id=assignment.id)

        except ClassSubject.DoesNotExist:
            messages.error(request, "Selected class and subject combination does not exist.")

    context = {
        'class_subjects': class_subjects,
    }
    return render(request, 'assignments/create_assignment.html', context)

@login_required
def assignment_detail(request, assignment_id):
    """
    Display details of an assignment including:
    - Basic information
    - Questions (if it's a quiz)
    - Submission status (for teachers)
    - Student's submission status (for students)
    """
    assignment = get_object_or_404(Assignment, id=assignment_id)

    # Common context data
    context = {
        'assignment': assignment,
        'current_time': timezone.now(),
    }

    # If it's a quiz, load questions
    if assignment.assignment_type == 'QUIZ':
        questions = Question.objects.filter(assignment=assignment).order_by('order')
        context['questions'] = questions

    # Data for teachers and admins (submission statistics)
    if is_teacher(request.user) or is_admin(request.user):
        # Get all students in the class
        students = Student.objects.filter(enrolled_subjects=assignment.class_subject)
        student_count = students.count()

        # Get submissions
        submissions = StudentSubmission.objects.filter(assignment=assignment)
        submission_count = submissions.count()
        graded_count = submissions.filter(is_graded=True).count()
        pending_count = submission_count - graded_count

        # Calculate percentages
        submitted_percent = 0
        if student_count > 0:
            submitted_percent = round((submission_count / student_count) * 100)

        context.update({
            'student_count': student_count,
            'submission_count': submission_count,
            'graded_count': graded_count,
            'pending_count': pending_count,
            'submitted_percent': submitted_percent,
        })

    # Data for students (their submission)
    elif is_student(request.user):
        try:
            student = Student.objects.get(user=request.user)
            student_submission = StudentSubmission.objects.filter(
                student=student,
                assignment=assignment
            ).first()

            context['student_submission'] = student_submission
        except Student.DoesNotExist:
            pass

    return render(request, 'assignments/assignment_detail.html', context)

@login_required
@user_passes_test(lambda u: is_teacher(u) or is_admin(u))
def edit_assignment(request, assignment_id):
    """
    Edit an existing assignment. Only teachers who created the assignment or admins can edit.
    """
    # Get the assignment
    assignment = get_object_or_404(Assignment, id=assignment_id)

    # Security check - only the teacher who created the assignment or admins can edit
    if not (is_admin(request.user) or (is_teacher(request.user) and assignment.created_by == request.user)):
        messages.error(request, "You don't have permission to edit this assignment.")
        return redirect('assignments:assignment_detail', assignment_id=assignment.id)

    # Get the classes that the teacher is teaching
    class_subjects = []

    if is_teacher(request.user):
        try:
            teacher = Teacher.objects.get(user=request.user)
            # Get all class subjects this teacher is assigned to teach
            # This includes subjects where they are either:
            # 1. The class teacher of the classroom, or
            # 2. The subject teacher assigned to the class subject
            class_subjects = ClassSubject.objects.filter(
                Q(classroom__class_teacher=teacher) |
                Q(teacher=teacher)
            ).distinct()
        except Teacher.DoesNotExist:
            messages.error(request, "Teacher profile not found.")
            return redirect('dashboard:index')
    elif is_admin(request.user):
        class_subjects = ClassSubject.objects.all()

    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        class_subject_id = request.POST.get('class_subject')
        assignment_type = request.POST.get('assignment_type')
        max_score = request.POST.get('max_score')
        due_date = request.POST.get('due_date')
        due_time = request.POST.get('due_time')
        is_active = 'is_active' in request.POST

        # Combine date and time
        from datetime import datetime
        due_datetime_str = f"{due_date} {due_time}"
        due_datetime = datetime.strptime(due_datetime_str, "%Y-%m-%d %H:%M")

        # Validate inputs
        if not (title and description and class_subject_id and assignment_type and max_score and due_date and due_time):
            messages.error(request, "Please fill in all required fields.")
            return render(request, 'assignments/edit_assignment.html', {
                'assignment': assignment,
                'class_subjects': class_subjects,
            })

        try:
            class_subject = ClassSubject.objects.get(id=class_subject_id)

            # Get time limit for quizzes
            time_limit = None
            if assignment_type == 'QUIZ':
                time_limit = request.POST.get('time_limit')
                if time_limit and time_limit.strip():
                    time_limit = int(time_limit)
                else:
                    time_limit = None

            # Update assignment
            assignment.title = title
            assignment.description = description
            assignment.class_subject = class_subject
            assignment.assignment_type = assignment_type
            assignment.max_score = max_score
            assignment.due_date = due_datetime
            assignment.is_active = is_active
            assignment.time_limit = time_limit

            # Handle file attachment if provided
            if 'file_attachment' in request.FILES:
                # Delete old file if it exists
                if assignment.file_attachment:
                    assignment.file_attachment.delete()
                # Add new file
                assignment.file_attachment = request.FILES['file_attachment']

            # Handle file removal if requested
            if 'remove_file' in request.POST and assignment.file_attachment:
                assignment.file_attachment.delete()
                assignment.file_attachment = None

            assignment.save()

            messages.success(request, f"Assignment '{title}' updated successfully.")
            return redirect('assignments:assignment_detail', assignment_id=assignment.id)

        except ClassSubject.DoesNotExist:
            messages.error(request, "Selected class and subject combination does not exist.")

    # Split datetime into date and time components for form
    from datetime import datetime
    due_date = assignment.due_date.strftime('%Y-%m-%d')
    due_time = assignment.due_date.strftime('%H:%M')

    context = {
        'assignment': assignment,
        'class_subjects': class_subjects,
        'due_date': due_date,
        'due_time': due_time,
    }
    return render(request, 'assignments/edit_assignment.html', context)

@login_required
def delete_assignment(request, assignment_id):
    assignment = get_object_or_404(Assignment, id=assignment_id)

    # Check permissions - Allow admin or the teacher who created the assignment
    if not (request.user.is_staff or assignment.created_by == request.user):
        messages.error(request, 'You do not have permission to delete this assignment.')
        return redirect('assignments:assignment_list')

    if request.method == 'POST':
        assignment.delete()
        messages.success(request, f'Assignment "{assignment.title}" was successfully deleted.')
        return redirect('assignments:assignment_list')

    return redirect('assignments:assignment_list')

# Question views
@login_required
def question_list(request, assignment_id):
    return render(request, 'assignments/question_list.html')

@login_required
@user_passes_test(lambda u: is_teacher(u) or is_admin(u))
def create_question(request, assignment_id):
    """
    Create a new question for a quiz assignment.
    Only teachers who own the assignment or admins can create questions.
    """
    assignment = get_object_or_404(Assignment, id=assignment_id)

    # Security check - only the teacher who created the assignment or admins can add questions
    if not (is_admin(request.user) or (is_teacher(request.user) and assignment.created_by == request.user)):
        messages.error(request, "You don't have permission to add questions to this assignment.")
        return redirect('assignments:assignment_detail', assignment_id=assignment.id)

    # Check if this is a quiz-type assignment
    if assignment.assignment_type != 'QUIZ':
        messages.error(request, "Questions can only be added to quiz-type assignments.")
        return redirect('assignments:assignment_detail', assignment_id=assignment.id)

    # Get existing questions for the sidebar
    current_questions = Question.objects.filter(assignment=assignment).order_by('order')

    if request.method == 'POST':
        # Get form data
        question_text = request.POST.get('question_text')
        question_type = request.POST.get('question_type')
        points = request.POST.get('points')
        show_feedback = 'show_feedback' in request.POST

        # Validate inputs
        if not (question_text and question_type and points):
            messages.error(request, "Please fill in all required fields.")
            return render(request, 'assignments/create_question.html', {
                'assignment': assignment,
                'current_questions': current_questions
            })

        # Determine the order for the new question
        order = current_questions.count() + 1

        # Create the question
        question = Question.objects.create(
            assignment=assignment,
            question_text=question_text,
            question_type=question_type,
            points=points,
            show_feedback=show_feedback,
            order=order
        )

        # Handle question type specific data
        if question_type == 'MCQ':
            # Get multiple choice options
            choice_texts = request.POST.getlist('choice_text[]')
            correct_choice_index = request.POST.get('correct_choice')

            # Validate
            if not (choice_texts and correct_choice_index is not None):
                question.delete()  # Delete the question if validation fails
                messages.error(request, "Please provide choices and select the correct answer.")
                return render(request, 'assignments/create_question.html', {
                    'assignment': assignment,
                    'current_questions': current_questions
                })

            # Create choices
            for i, choice_text in enumerate(choice_texts):
                is_correct = (str(i) == str(correct_choice_index))

                Choice.objects.create(
                    question=question,
                    choice_text=choice_text,
                    is_correct=is_correct,
                    order=i
                )

        elif question_type == 'SHORT':
            # Store expected answer for reference (optional)
            expected_answer = request.POST.get('expected_answer')
            if expected_answer:
                # Store as a note in the question
                question.notes = f"Expected Answer: {expected_answer}"
                question.save()

        elif question_type == 'LONG':
            # Store grading notes (optional)
            grading_notes = request.POST.get('grading_notes')
            if grading_notes:
                # Store as a note in the question
                question.notes = f"Grading Notes: {grading_notes}"
                question.save()

        elif question_type == 'FILE':
            # Store allowed file types (optional)
            allowed_file_types = request.POST.get('allowed_file_types')
            if allowed_file_types:
                # Store as a note in the question
                question.notes = f"Allowed File Types: {allowed_file_types}"
                question.save()

        messages.success(request, "Question added successfully.")

        # Check if "Save & Add Another" was clicked
        if 'save_and_add' in request.POST:
            return redirect('assignments:create_question', assignment_id=assignment.id)
        else:
            return redirect('assignments:assignment_detail', assignment_id=assignment.id)

    # GET request - display the form
    context = {
        'assignment': assignment,
        'current_questions': current_questions
    }

    return render(request, 'assignments/create_question.html', context)

@login_required
def question_detail(request, question_id):
    """
    Display detailed information about a specific question.
    Shows question text, type, choices (for MCQs), and other details.
    Includes navigation to previous and next questions.
    """
    # Get the question
    question = get_object_or_404(Question, id=question_id)
    assignment = question.assignment

    # Check permissions - anyone with access to the assignment can view questions
    if is_student(request.user):
        # Students can only view questions if they are enrolled in the class
        student = get_object_or_404(Student, user=request.user)
        if not student.enrolled_subjects.filter(id=assignment.class_subject.id).exists():
            messages.error(request, "You don't have permission to view this question.")
            return redirect('assignments:assignment_list')
    elif is_teacher(request.user):
        # Subject teachers can only view questions for their assigned subjects
        teacher = get_object_or_404(Teacher, user=request.user)
        if not (
            assignment.created_by == request.user or
            ClassSubject.objects.filter(teacher=teacher, id=assignment.class_subject.id).exists() or
            ClassRoom.objects.filter(class_teacher=teacher, id=assignment.class_subject.classroom.id).exists()
        ):
            messages.error(request, "You don't have permission to view this question.")
            return redirect('assignments:assignment_list')
    elif is_parent(request.user):
        # Parents can only view questions for their children's classes
        parent = get_object_or_404(Parent, user=request.user)
        if not parent.children.filter(enrolled_subjects=assignment.class_subject).exists():
            messages.error(request, "You don't have permission to view this question.")
            return redirect('assignments:assignment_list')
    elif not is_admin(request.user):
        # If not a student, teacher, parent, or admin, deny access
        messages.error(request, "You don't have permission to view this question.")
        return redirect('assignments:assignment_list')

    # Get choices for MCQ questions
    choices = []
    if question.question_type == Question.QuestionType.MULTIPLE_CHOICE:
        choices = Choice.objects.filter(question=question).order_by('order')

    # Prepare extra data based on question type
    extra_data = {}
    if question.notes:
        if question.question_type == 'SHORT' and question.notes.startswith("Expected Answer:"):
            extra_data['expected_answer'] = question.notes.replace("Expected Answer:", "").strip()
        elif question.question_type == 'LONG' and question.notes.startswith("Grading Notes:"):
            extra_data['grading_notes'] = question.notes.replace("Grading Notes:", "").strip()
        elif question.question_type == 'FILE' and question.notes.startswith("Allowed File Types:"):
            extra_data['allowed_file_types'] = question.notes.replace("Allowed File Types:", "").strip()

    # Get previous and next questions for navigation
    prev_question = Question.objects.filter(
        assignment=assignment,
        order__lt=question.order
    ).order_by('-order').first()

    next_question = Question.objects.filter(
        assignment=assignment,
        order__gt=question.order
    ).order_by('order').first()

    # Prepare context
    context = {
        'question': question,
        'assignment': assignment,
        'choices': choices,
        'extra_data': extra_data,
        'prev_question': prev_question,
        'next_question': next_question,
    }

    return render(request, 'assignments/question_detail.html', context)

@login_required
@user_passes_test(lambda u: is_teacher(u) or is_admin(u))
def edit_question(request, question_id):
    """
    Edit an existing question for a quiz assignment.
    Supports all question types and updating choices for MCQs.
    """
    # Get the question
    question = get_object_or_404(Question, id=question_id)
    assignment = question.assignment

    # Security check - only the teacher who created the assignment or admins can edit questions
    if not (is_admin(request.user) or (is_teacher(request.user) and assignment.created_by == request.user)):
        messages.error(request, "You don't have permission to edit questions for this assignment.")
        return redirect('assignments:assignment_detail', assignment_id=assignment.id)

    # Get existing choices for MCQs
    choices = []
    if question.question_type == Question.QuestionType.MULTIPLE_CHOICE:
        choices = Choice.objects.filter(question=question).order_by('order')

    # Get existing questions for the sidebar
    current_questions = Question.objects.filter(assignment=assignment).order_by('order')

    if request.method == 'POST':
        # Get form data
        question_text = request.POST.get('question_text')
        question_type = request.POST.get('question_type')
        points = request.POST.get('points')
        show_feedback = 'show_feedback' in request.POST

        # Validate inputs
        if not (question_text and question_type and points):
            messages.error(request, "Please fill in all required fields.")
            return render(request, 'assignments/edit_question.html', {
                'question': question,
                'choices': choices,
                'assignment': assignment,
                'current_questions': current_questions
            })

        # Update the question
        question.question_text = question_text
        question.question_type = question_type
        question.points = points
        question.show_feedback = show_feedback

        # Handle question type specific data
        if question_type == 'MCQ':
            # Get multiple choice options
            choice_texts = request.POST.getlist('choice_text[]')
            choice_ids = request.POST.getlist('choice_id[]', [])
            correct_choice_index = request.POST.get('correct_choice')

            # Validate
            if not (choice_texts and correct_choice_index is not None):
                messages.error(request, "Please provide choices and select the correct answer.")
                return render(request, 'assignments/edit_question.html', {
                    'question': question,
                    'choices': choices,
                    'assignment': assignment,
                    'current_questions': current_questions
                })

            # Delete existing choices not in the updated list
            existing_choice_ids = [str(c.id) for c in choices]
            for choice_id in existing_choice_ids:
                if choice_id not in choice_ids:
                    try:
                        Choice.objects.get(id=choice_id).delete()
                    except Choice.DoesNotExist:
                        pass

            # Update or create choices
            for i, choice_text in enumerate(choice_texts):
                is_correct = (str(i) == str(correct_choice_index))

                # Check if this is an existing choice
                choice_id = choice_ids[i] if i < len(choice_ids) else None

                if choice_id and choice_id in existing_choice_ids:
                    # Update existing choice
                    try:
                        choice = Choice.objects.get(id=choice_id)
                        choice.choice_text = choice_text
                        choice.is_correct = is_correct
                        choice.order = i
                        choice.save()
                    except Choice.DoesNotExist:
                        # Create as new if choice_id doesn't exist
                        Choice.objects.create(
                            question=question,
                            choice_text=choice_text,
                            is_correct=is_correct,
                            order=i
                        )
                else:
                    # Create new choice
                    Choice.objects.create(
                        question=question,
                        choice_text=choice_text,
                        is_correct=is_correct,
                        order=i
                    )

        elif question_type == 'SHORT':
            # Store expected answer for reference (optional)
            expected_answer = request.POST.get('expected_answer')
            if expected_answer:
                # Store in notes field
                question.notes = f"Expected Answer: {expected_answer}"

        elif question_type == 'LONG':
            # Store grading notes (optional)
            grading_notes = request.POST.get('grading_notes')
            if grading_notes:
                # Store in notes field
                question.notes = f"Grading Notes: {grading_notes}"

        elif question_type == 'FILE':
            # Store allowed file types (optional)
            allowed_file_types = request.POST.get('allowed_file_types')
            if allowed_file_types:
                # Store in notes field
                question.notes = f"Allowed File Types: {allowed_file_types}"

        # Save the updated question
        question.save()

        messages.success(request, "Question updated successfully.")
        return redirect('assignments:assignment_detail', assignment_id=assignment.id)

    # Prepare data for the form
    notes_data = {}
    if question.notes:
        if question.question_type == 'SHORT' and question.notes.startswith("Expected Answer:"):
            notes_data['expected_answer'] = question.notes.replace("Expected Answer:", "").strip()
        elif question.question_type == 'LONG' and question.notes.startswith("Grading Notes:"):
            notes_data['grading_notes'] = question.notes.replace("Grading Notes:", "").strip()
        elif question.question_type == 'FILE' and question.notes.startswith("Allowed File Types:"):
            notes_data['allowed_file_types'] = question.notes.replace("Allowed File Types:", "").strip()

    # GET request - display the form
    context = {
        'question': question,
        'choices': choices,
        'assignment': assignment,
        'current_questions': current_questions,
        'notes_data': notes_data
    }

    return render(request, 'assignments/edit_question.html', context)

@login_required
@user_passes_test(lambda u: is_teacher(u) or is_admin(u))
def delete_question(request, question_id):
    """
    Delete a question from a quiz assignment.
    Requires confirmation to prevent accidental deletion.
    """
    # Get the question
    question = get_object_or_404(Question, id=question_id)
    assignment = question.assignment

    # Security check - only the teacher who created the assignment or admins can delete questions
    if not (is_admin(request.user) or (is_teacher(request.user) and assignment.created_by == request.user)):
        messages.error(request, "You don't have permission to delete questions from this assignment.")
        return redirect('assignments:assignment_detail', assignment_id=assignment.id)

    # Get existing choices for MCQs for display
    choices = []
    if question.question_type == Question.QuestionType.MULTIPLE_CHOICE:
        choices = Choice.objects.filter(question=question).order_by('order')

    if request.method == 'POST':
        # Delete choices first (to avoid foreign key constraint errors)
        if choices:
            for choice in choices:
                choice.delete()

        # Get the question order for reordering
        question_order = question.order

        # Delete the question
        question.delete()

        # Reorder remaining questions
        remaining_questions = Question.objects.filter(
            assignment=assignment,
            order__gt=question_order
        ).order_by('order')

        for i, q in enumerate(remaining_questions):
            q.order = question_order + i
            q.save()

        messages.success(request, "Question deleted successfully.")
        return redirect('assignments:assignment_detail', assignment_id=assignment.id)

    # Prepare context for confirmation page
    context = {
        'question': question,
        'choices': choices,
    }

    return render(request, 'assignments/delete_question.html', context)

# Choice views
@login_required
def choice_list(request, question_id):
    return render(request, 'assignments/choice_list.html')

@login_required
def create_choice(request, question_id):
    return render(request, 'assignments/create_choice.html')

@login_required
def edit_choice(request, choice_id):
    return render(request, 'assignments/edit_choice.html')

@login_required
def delete_choice(request, choice_id):
    return render(request, 'assignments/delete_choice.html')

# Student submission views
@login_required
@user_passes_test(lambda u: is_student(u))
def submit_assignment(request, assignment_id):
    """
    Allow students to submit assignments.
    - Can be text submissions, file uploads, or both
    - Handles updating existing submissions
    """
    # Get the assignment
    assignment = get_object_or_404(Assignment, id=assignment_id)

    # Check if the student is enrolled in this class
    try:
        student = Student.objects.get(user=request.user)

        # Verify student is enrolled in the class
        if not student.enrolled_subjects.filter(id=assignment.class_subject.id).exists():
            messages.error(request, "You are not enrolled in this class.")
            return redirect('assignments:assignment_list')

        # Check if student has already submitted
        existing_submission = StudentSubmission.objects.filter(
            student=student,
            assignment=assignment
        ).first()

        # Check if deadline has passed
        deadline_passed = timezone.now() > assignment.due_date

        if request.method == 'POST':
            text_submission = request.POST.get('text_submission', '').strip()
            remove_file = 'remove_file' in request.POST

            # Check if there's any submission content
            has_file = 'file_submission' in request.FILES and request.FILES['file_submission']

            if not text_submission and not has_file and not existing_submission:
                messages.error(request, "Please provide a text response or upload a file.")
                return render(request, 'assignments/submit_assignment.html', {
                    'assignment': assignment,
                    'existing_submission': existing_submission,
                    'deadline_passed': deadline_passed
                })

            # Create or update submission
            if existing_submission:
                submission = existing_submission
                submission.text_submission = text_submission

                # Handle file submission
                if has_file:
                    submission.file_submission = request.FILES['file_submission']
                elif remove_file and submission.file_submission:
                    # Delete the file if remove_file is checked
                    submission.file_submission.delete()
                    submission.file_submission = None

                submission.submission_date = timezone.now()
                submission.is_graded = False  # Reset graded status on resubmission
                submission.save()

                messages.success(request, "Your assignment has been updated successfully.")
            else:
                # Create new submission
                submission = StudentSubmission.objects.create(
                    student=student,
                    assignment=assignment,
                    text_submission=text_submission,
                    submission_date=timezone.now()
                )

                # Handle file submission
                if has_file:
                    submission.file_submission = request.FILES['file_submission']
                    submission.save()

                messages.success(request, "Your assignment has been submitted successfully.")

            return redirect('assignments:assignment_detail', assignment_id=assignment.id)

        # Display the submission form
        context = {
            'assignment': assignment,
            'existing_submission': existing_submission,
            'deadline_passed': deadline_passed
        }

        return render(request, 'assignments/submit_assignment.html', context)

    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('dashboard:index')

@login_required
@user_passes_test(lambda u: is_teacher(u) or is_admin(u))
def submission_list(request):
    """
    Display a list of submissions for teachers to review and grade.
    Allows filtering by assignment, status, etc.
    """
    # Get filter parameters
    assignment_id = request.GET.get('assignment_id')
    status_filter = request.GET.get('status', 'all')

    # Initialize variables for template context
    filter_assignment = None
    submissions = []

    # Get the appropriate submissions based on user role
    if is_teacher(request.user):
        try:
            teacher = Teacher.objects.get(user=request.user)
            # Check if this is a subject teacher
            is_subject_teacher = not ClassRoom.objects.filter(class_teacher=teacher).exists()

            if is_subject_teacher:
                # Subject teachers see submissions for their assigned subjects only
                if assignment_id:
                    try:
                        filter_assignment = Assignment.objects.get(id=assignment_id, class_subject__teacher=teacher)
                        submissions = StudentSubmission.objects.filter(assignment=filter_assignment)
                    except Assignment.DoesNotExist:
                        messages.error(request, "Assignment not found or you don't have permission to view it.")
                        return redirect('assignments:assignment_list')
                else:
                    submissions = StudentSubmission.objects.filter(assignment__class_subject__teacher=teacher)
            else:
                # Class teachers see all submissions for their assigned class
                class_teacher_of = ClassRoom.objects.filter(class_teacher=teacher)
                if assignment_id:
                    try:
                        filter_assignment = Assignment.objects.get(id=assignment_id, class_subject__classroom__in=class_teacher_of)
                        submissions = StudentSubmission.objects.filter(assignment=filter_assignment)
                    except Assignment.DoesNotExist:
                         messages.error(request, "Assignment not found or you don't have permission to view it.")
                         return redirect('assignments:assignment_list')
                else:
                    submissions = StudentSubmission.objects.filter(assignment__class_subject__classroom__in=class_teacher_of)
        except Teacher.DoesNotExist:
            submissions = []  # Handle case where teacher profile doesn't exist

    elif is_admin(request.user):
        # Admins can see all submissions
        if assignment_id:
            try:
                filter_assignment = Assignment.objects.get(id=assignment_id)
                submissions = StudentSubmission.objects.filter(assignment=filter_assignment)
            except Assignment.DoesNotExist:
                messages.error(request, "Assignment not found.")
                return redirect('assignments:assignment_list')
        else:
            submissions = StudentSubmission.objects.all()

    # Apply status filter
    if status_filter != 'all':
        if status_filter == 'graded':
            submissions = submissions.filter(is_graded=True)
        elif status_filter == 'pending':
            submissions = submissions.filter(is_graded=False)
        elif status_filter == 'ontime':
            # Submissions made before due date
            submissions = submissions.filter(submission_date__lte=F('assignment__due_date'))
        elif status_filter == 'late':
            # Submissions made after due date
            submissions = submissions.filter(submission_date__gt=F('assignment__due_date'))

    # Calculate statistics if filtering by a specific assignment
    if filter_assignment:
        # Get all students in the class
        students = Student.objects.filter(enrolled_subjects=filter_assignment.class_subject)
        student_count = students.count()

        # Submission statistics
        submission_count = submissions.count()
        graded_count = submissions.filter(is_graded=True).count()
        pending_count = submission_count - graded_count
        late_count = submissions.filter(submission_date__gt=F('assignment__due_date')).count()

        # Calculate percentages
        submission_percent = 0
        if student_count > 0:
            submission_percent = round((submission_count / student_count) * 100)

        # Find students who haven't submitted
        submitted_student_ids = submissions.values_list('student_id', flat=True)
        not_submitted_students = students.exclude(id__in=submitted_student_ids)
        not_submitted = not_submitted_students.count()

        # Grade distribution if there are graded submissions
        grade_ranges = ['90-100%', '80-89%', '70-79%', '60-69%', 'Below 60%']
        grade_distribution = [0, 0, 0, 0, 0]
        average_score = 0
        highest_score = 0
        lowest_score = filter_assignment.max_score

        graded_submissions = submissions.filter(is_graded=True)
        if graded_submissions.exists():
            # Calculate grade statistics
            scores = [sub.score for sub in graded_submissions]
            if scores:
                average_score = sum(scores) / len(scores)
                highest_score = max(scores)
                lowest_score = min(scores)

                # Calculate grade distribution
                max_possible = filter_assignment.max_score
                for sub in graded_submissions:
                    percent = (sub.score / max_possible) * 100
                    if percent >= 90:
                        grade_distribution[0] += 1
                    elif percent >= 80:
                        grade_distribution[1] += 1
                    elif percent >= 70:
                        grade_distribution[2] += 1
                    elif percent >= 60:
                        grade_distribution[3] += 1
                    else:
                        grade_distribution[4] += 1

        # Add all statistics to context
        context = {
            'submissions': submissions,
            'filter_assignment': filter_assignment,
            'status_filter': status_filter,
            'student_count': student_count,
            'submission_count': submission_count,
            'graded_count': graded_count,
            'pending_count': pending_count,
            'late_count': late_count,
            'submission_percent': submission_percent,
            'not_submitted': not_submitted,
            'not_submitted_students': not_submitted_students,
            'grade_ranges': grade_ranges,
            'grade_distribution': grade_distribution,
            'average_score': average_score,
            'highest_score': highest_score,
            'lowest_score': lowest_score,
        }
    else:
        # Basic context for all submissions view
        context = {
            'submissions': submissions,
            'status_filter': status_filter,
        }

    return render(request, 'assignments/submission_list.html', context)

@login_required
def submission_detail(request, submission_id):
    """
    Display detailed information about a specific submission.
    Shows the student's answers, score, and teacher feedback.
    """
    # Get the submission
    submission = get_object_or_404(StudentSubmission, id=submission_id)

    # Check permissions based on user role
    user = request.user

    if is_student(user):
        # Students can only view their own submissions
        if submission.student.user != user:
            messages.error(request, "You don't have permission to view this submission.")
            return redirect('assignments:assignment_list')
    elif is_parent(user):
        # Parents can only view their children's submissions
        parent = get_object_or_404(Parent, user=user)
        if not parent.children.filter(id=submission.student.id).exists():
            messages.error(request, "You don't have permission to view this submission.")
            return redirect('assignments:assignment_list')
    elif is_teacher(user):
        # Teachers can only view submissions for their classes
        teacher = get_object_or_404(Teacher, user=user)
        classroom = submission.assignment.class_subject.classroom
        class_subject = submission.assignment.class_subject

        # Check if teacher is assigned to this class or subject
        if not (ClassRoom.objects.filter(class_teacher=teacher, id=classroom.id).exists() or
                class_subject.teacher == teacher):
            messages.error(request, "You don't have permission to view this submission.")
            return redirect('assignments:assignment_list')
    elif not is_admin(user):
        # If not a student, parent, teacher, or admin, deny access
        messages.error(request, "You don't have permission to view this submission.")
        return redirect('assignments:assignment_list')

    # Prepare context
    context = {
        'submission': submission,
    }

    # If this is a quiz submission, get the student's answers
    if submission.assignment.assignment_type == 'QUIZ':
        student_answers = StudentAnswer.objects.filter(
            submission=submission
        ).order_by('question__order')

        context['student_answers'] = student_answers

    return render(request, 'assignments/submission_detail.html', context)

@login_required
@user_passes_test(lambda u: is_teacher(u) or is_admin(u))
def view_submission(request, submission_id):
    """
    Display a student's submission for review by a teacher.
    Shows text responses, file uploads, and for quizzes shows all answers.
    """
    # Get the submission
    submission = get_object_or_404(StudentSubmission, id=submission_id)

    # Security check for teachers - verify appropriate permissions
    if is_teacher(request.user):
        try:
            teacher = Teacher.objects.get(user=request.user)

            # Check if this is a subject teacher (not a class teacher for any class)
            is_subject_teacher = not ClassRoom.objects.filter(class_teacher=teacher).exists()

            if is_subject_teacher:
                # Subject teachers can only view submissions for their assigned subjects
                if not ClassSubject.objects.filter(
                    teacher=teacher,
                    id=submission.assignment.class_subject.id
                ).exists():
                    messages.error(request, "You don't have permission to view this submission.")
                    return redirect('assignments:assignment_list')
            else:
                # Class teachers can view submissions for their assigned classes
                classroom = submission.assignment.class_subject.classroom
                if not ClassRoom.objects.filter(class_teacher=teacher, id=classroom.id).exists():
                    messages.error(request, "You don't have permission to view this submission.")
                    return redirect('assignments:assignment_list')
        except Teacher.DoesNotExist:
            messages.error(request, "Teacher profile not found.")
            return redirect('assignments:assignment_list')

    # Initialize context
    context = {
        'submission': submission,
    }

    # If this is a quiz submission, get the student's answers
    if submission.assignment.assignment_type == 'QUIZ':
        student_answers = StudentAnswer.objects.filter(
            submission=submission
        ).order_by('question__order')

        context['student_answers'] = student_answers

    return render(request, 'assignments/view_submission.html', context)

@login_required
@user_passes_test(lambda u: is_teacher(u) or is_admin(u))
def grade_submission(request, submission_id):
    """
    Allow teachers to grade student submissions.
    Handles both regular assignments and quizzes with different grading approaches.
    """
    # Get the submission
    submission = get_object_or_404(StudentSubmission, id=submission_id)

    # Security check for teachers - verify appropriate permissions
    if is_teacher(request.user):
        try:
            teacher = Teacher.objects.get(user=request.user)

            # Check if this is a subject teacher (not a class teacher for any class)
            is_subject_teacher = not ClassRoom.objects.filter(class_teacher=teacher).exists()

            if is_subject_teacher:
                # Subject teachers can only grade submissions for their assigned subjects
                if not ClassSubject.objects.filter(
                    teacher=teacher,
                    id=submission.assignment.class_subject.id
                ).exists():
                    messages.error(request, "You don't have permission to grade this submission.")
                    return redirect('assignments:assignment_list')
            else:
                # Class teachers can grade submissions for their assigned classes
                classroom = submission.assignment.class_subject.classroom
                if not ClassRoom.objects.filter(class_teacher=teacher, id=classroom.id).exists():
                    messages.error(request, "You don't have permission to grade this submission.")
                    return redirect('assignments:assignment_list')
        except Teacher.DoesNotExist:
            messages.error(request, "Teacher profile not found.")
            return redirect('assignments:assignment_list')

    # Initialize context
    context = {
        'submission': submission,
    }

    # Handle POST request for grade submission
    if request.method == 'POST':
        # Get overall score and comment
        score = request.POST.get('score')
        teacher_comment = request.POST.get('teacher_comment')

        # Validate score - ensure it's within valid range
        try:
            score = float(score)
            if score < 0 or score > submission.assignment.max_score:
                messages.error(request, f"Score must be between 0 and {submission.assignment.max_score}.")
                return redirect('assignments:grade_submission', submission_id=submission.id)
        except (ValueError, TypeError):
            messages.error(request, "Invalid score value.")
            return redirect('assignments:grade_submission', submission_id=submission.id)

        # Update submission with grade
        submission.score = score
        submission.teacher_comment = teacher_comment
        submission.is_graded = True
        submission.save()

        # If it's a quiz, handle individual question grades
        if submission.assignment.assignment_type == 'QUIZ':
            # Get all student answers for this submission
            student_answers = StudentAnswer.objects.filter(submission=submission)

            # Process each answer's grade
            for answer in student_answers:
                # Get points and feedback for this answer
                points_key = f'points_{answer.id}'
                feedback_key = f'feedback_{answer.id}'
                is_correct_key = f'is_correct_{answer.id}'

                if points_key in request.POST:
                    try:
                        points = float(request.POST.get(points_key))
                        # Ensure points are within valid range
                        if points < 0 or points > answer.question.points:
                            points = 0
                    except (ValueError, TypeError):
                        points = 0

                    # Determine correctness
                    is_correct_val = request.POST.get(is_correct_key, 'incorrect')
                    is_correct = is_correct_val == 'correct'

                    # Get feedback
                    feedback = request.POST.get(feedback_key, '')

                    # Update the answer
                    answer.score = points
                    answer.is_correct = is_correct
                    answer.feedback = feedback
                    answer.save()

        messages.success(request, "Submission graded successfully.")
        return redirect('assignments:view_submission', submission_id=submission.id)

    # If it's a quiz, get the student's answers
    if submission.assignment.assignment_type == 'QUIZ':
        student_answers = StudentAnswer.objects.filter(
            submission=submission
        ).order_by('question__order')

        context['student_answers'] = student_answers

    return render(request, 'assignments/grade_submission.html', context)

@login_required
@user_passes_test(lambda u: is_teacher(u) or is_admin(u))
def auto_grade_submission(request, submission_id):
    """
    Automatically grade a quiz submission by comparing student
    answers with correct answers.
    Only works for multiple choice questions.
    """
    # Get the submission
    submission = get_object_or_404(StudentSubmission, id=submission_id)

    # Security check for teachers - verify appropriate permissions
    if is_teacher(request.user):
        try:
            teacher = Teacher.objects.get(user=request.user)

            # Check if this is a subject teacher (not a class teacher for any class)
            is_subject_teacher = not ClassRoom.objects.filter(class_teacher=teacher).exists()

            if is_subject_teacher:
                # Subject teachers can only auto-grade submissions for their assigned subjects
                if not ClassSubject.objects.filter(
                    teacher=teacher,
                    id=submission.assignment.class_subject.id
                ).exists():
                    messages.error(request, "You don't have permission to auto-grade this submission.")
                    return redirect('assignments:assignment_list')
            else:
                # Class teachers can auto-grade submissions for their assigned classes
                classroom = submission.assignment.class_subject.classroom
                if not ClassRoom.objects.filter(class_teacher=teacher, id=classroom.id).exists():
                    messages.error(request, "You don't have permission to auto-grade this submission.")
                    return redirect('assignments:assignment_list')
        except Teacher.DoesNotExist:
            messages.error(request, "Teacher profile not found.")
            return redirect('assignments:assignment_list')

    # Check if this is a quiz
    if submission.assignment.assignment_type != 'QUIZ':
        messages.error(request, "Auto-grading only works for quizzes.")
        return redirect('assignments:view_submission', submission_id=submission.id)

    # Get all student answers for this submission
    student_answers = StudentAnswer.objects.filter(
        submission=submission
    ).select_related('question', 'selected_choice')

    if not student_answers.exists():
        messages.error(request, "No answers found for this submission.")
        return redirect('assignments:view_submission', submission_id=submission.id)

    # Track grading stats
    total_questions = 0
    correct_answers = 0
    total_points = 0
    max_points = 0

    # Process each answer
    for answer in student_answers:
        question = answer.question
        max_points += float(question.points)
        total_questions += 1

        # Only auto-grade MCQ questions
        if question.question_type == 'MCQ':
            # Get the correct choice
            try:
                correct_choice = Choice.objects.get(question=question, is_correct=True)

                # Compare with student's choice
                if answer.selected_choice and answer.selected_choice.id == correct_choice.id:
                    # Correct answer
                    answer.is_correct = True
                    answer.score = question.points
                    total_points += float(question.points)
                    correct_answers += 1

                    if question.show_feedback:
                        answer.feedback = "Correct answer!"
                else:
                    # Incorrect answer
                    answer.is_correct = False
                    answer.score = 0

                    if question.show_feedback:
                        answer.feedback = f"Incorrect. The correct answer is: {correct_choice.choice_text}"

                answer.save()

            except Choice.DoesNotExist:
                # Skip questions with no correct answer defined
                continue

    # Update submission with overall score
    if max_points > 0:
        submission.score = total_points
        submission.is_graded = True
        submission.teacher_comment = (f"Auto-graded on {timezone.now().strftime('%Y-%m-%d %H:%M')}. "
                                      f"Score: {correct_answers}/{total_questions} questions correct.")
        submission.save()

        messages.success(request,
                        f"Quiz auto-graded successfully. Score: {submission.score}/{submission.assignment.max_score}")
    else:
        messages.warning(request, "No multiple-choice questions could be auto-graded.")

    return redirect('assignments:view_submission', submission_id=submission.id)

# Grading views
@login_required
def grade_list(request):
    return render(request, 'assignments/grade_list.html')

@login_required
def create_grade(request):
    return render(request, 'assignments/create_grade.html')

@login_required
def grade_detail(request, grade_id):
    return render(request, 'assignments/grade_detail.html')

@login_required
def edit_grade(request, grade_id):
    return render(request, 'assignments/edit_grade.html')

@login_required
def delete_grade(request, grade_id):
    return render(request, 'assignments/delete_grade.html')

# Quiz views with MCQ support
@login_required
def quiz_list(request):
    return render(request, 'assignments/quiz_list.html')

@login_required
def create_quiz(request):
    return render(request, 'assignments/create_quiz.html')

@login_required
def quiz_detail(request, quiz_id):
    return render(request, 'assignments/quiz_detail.html')

@login_required
def edit_quiz(request, quiz_id):
    return render(request, 'assignments/edit_quiz.html')

@login_required
def delete_quiz(request, quiz_id):
    return render(request, 'assignments/delete_quiz.html')

# Student quiz taking views
@login_required
@user_passes_test(lambda u: is_student(u))
def take_quiz(request, assignment_id):
    """
    Allow students to take a quiz assignment.
    - Display all questions (MCQ, short answer, long answer, file upload)
    - Process and save answers
    - Track time remaining
    """
    # Get the assignment
    assignment = get_object_or_404(Assignment, id=assignment_id)

    # Verify this is a quiz assignment
    if assignment.assignment_type != 'QUIZ':
        messages.error(request, "This is not a quiz assignment.")
        return redirect('assignments:assignment_detail', assignment_id=assignment.id)

    # Check if the student is enrolled in this class
    try:
        student = Student.objects.get(user=request.user)

        # Verify student is enrolled in the class
        if not student.enrolled_subjects.filter(id=assignment.class_subject.id).exists():
            messages.error(request, "You are not enrolled in this class.")
            return redirect('assignments:assignment_list')

        # Check if student has already submitted
        existing_submission = StudentSubmission.objects.filter(
            student=student,
            assignment=assignment
        ).first()

        # Get questions
        all_questions = list(Question.objects.filter(assignment=assignment).order_by('order'))

        # Calculate total points for the assignment's max_score
        total_quiz_points = sum(q.points for q in all_questions)

        # Implement question bank functionality - if questions_to_display is set and less than total questions
        if assignment.questions_to_display and len(all_questions) > assignment.questions_to_display:
            import random
            # Determine which questions to show
            if assignment.randomize_questions:
                # Randomly select the specified number of questions
                selected_questions = random.sample(all_questions, assignment.questions_to_display)
                # Sort selected questions by their original order
                selected_questions.sort(key=lambda q: q.order)
            else:
                # If randomization is disabled but question bank is enabled,
                # take the first N questions in original order
                selected_questions = all_questions[:assignment.questions_to_display]

            # Calculate adjusted max_score based on selected questions
            adjusted_max_score = sum(q.points for q in selected_questions)
        else:
            selected_questions = all_questions
            adjusted_max_score = total_quiz_points

        # For each MCQ question, shuffle its choices only if randomize_choices is enabled
        questions = []
        for question in selected_questions:
            if question.question_type == 'MCQ':
                choices = list(Choice.objects.filter(question=question))
                if assignment.randomize_choices:
                    # Shuffle the choices
                    import random
                    random.shuffle(choices)
                # Attach choices to the question (shuffled or original order)
                question.shuffled_choices = choices
            questions.append(question)

        # Calculate time remaining until due date
        now = timezone.now()
        time_remaining = assignment.due_date - now

        # Process quiz submission
        if request.method == 'POST' and not existing_submission and (not time_remaining.days < 0):
            # Create submission record
            submission = StudentSubmission.objects.create(
                student=student,
                assignment=assignment,
                submission_date=now
            )

            # Process answers for each question
            for question in questions:
                answer = None

                # Handle different question types
                if question.question_type == 'MCQ':
                    # Multiple choice
                    choice_id = request.POST.get(f'choice_{question.id}')
                    if choice_id:
                        try:
                            selected_choice = Choice.objects.get(id=choice_id, question=question)
                            answer = StudentAnswer.objects.create(
                                submission=submission,
                                question=question,
                                selected_choice=selected_choice
                            )
                        except Choice.DoesNotExist:
                            # Invalid choice, create empty answer
                            answer = StudentAnswer.objects.create(
                                submission=submission,
                                question=question
                            )
                    else:
                        # No choice selected, create empty answer
                        answer = StudentAnswer.objects.create(
                            submission=submission,
                            question=question
                        )

                elif question.question_type in ['SHORT', 'LONG']:
                    # Text answers
                    text_answer = request.POST.get(f'text_{question.id}', '').strip()
                    answer = StudentAnswer.objects.create(
                        submission=submission,
                        question=question,
                        text_answer=text_answer
                    )

                elif question.question_type == 'FILE':
                    # File upload
                    file_field = f'file_{question.id}'
                    if file_field in request.FILES:
                        answer = StudentAnswer.objects.create(
                            submission=submission,
                            question=question,
                            file_answer=request.FILES[file_field]
                        )
                    else:
                        # No file uploaded, create empty answer
                        answer = StudentAnswer.objects.create(
                            submission=submission,
                            question=question
                        )

            # Auto-grade MCQ questions immediately
            # Get all MCQ questions with answers
            mcq_answers = StudentAnswer.objects.filter(
                submission=submission,
                question__question_type='MCQ'
            )

            # Track auto-grading stats
            auto_graded_count = 0
            auto_correct_count = 0
            total_points = 0

            # Auto-grade each MCQ answer
            for answer in mcq_answers:
                # Get the correct choice for this MCQ question
                try:
                    correct_choice = Choice.objects.get(question=answer.question, is_correct=True)

                    # Compare with student's answer
                    if answer.selected_choice and answer.selected_choice.id == correct_choice.id:
                        # Correct answer
                        answer.is_correct = True
                        answer.points_earned = answer.question.points
                        answer.score = answer.question.points  # Set score for backwards compatibility
                        total_points += float(answer.question.points)
                        auto_correct_count += 1

                        if answer.question.show_feedback:
                            answer.feedback = "Correct answer!"
                    else:
                        # Incorrect answer
                        answer.is_correct = False
                        answer.points_earned = 0
                        answer.score = 0

                        if answer.question.show_feedback:
                            answer.feedback = f"Incorrect. The correct answer is: {correct_choice.choice_text}"

                    answer.save()
                    auto_graded_count += 1

                except Choice.DoesNotExist:
                    # Skip questions with no correct answer defined
                    continue

            # If we have auto-graded some questions, update the submission
            if auto_graded_count > 0:
                # Check if all questions were MCQ and auto-graded
                total_questions = StudentAnswer.objects.filter(submission=submission).count()

                if auto_graded_count == total_questions:
                    # All questions were auto-graded, mark the submission as graded
                    submission.is_graded = True
                    # Use the adjusted max score for this set of questions
                    submission.score = total_points
                    submission.save()

            # Only add one success message
            messages.success(request, "Quiz submitted successfully!")

            return redirect('assignments:quiz_result', assignment_id=assignment.id)

        # Prepare context for template
        context = {
            'assignment': assignment,
            'questions': questions,
            'existing_submission': existing_submission,
            'time_remaining': time_remaining,
        }

        return render(request, 'assignments/take_quiz.html', context)

    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('dashboard:index')

@login_required
@user_passes_test(lambda u: is_student(u))
def retake_quiz(request, assignment_id):
    """
    Allow students to retake a previously submitted quiz with enhanced security measures.
    - Uses QuizAttempt model to track and limit attempts
    - Implements age-appropriate security controls based on grade level
    - Prevents exploitation of the quiz system with cooldown periods
    - Randomizes questions for older students to prevent memorization
    """
    # Get the assignment
    assignment = get_object_or_404(Assignment, id=assignment_id)

    # Verify this is a quiz assignment
    if assignment.assignment_type != 'QUIZ':
        messages.error(request, "This is not a quiz assignment.")
        return redirect('assignments:assignment_detail', assignment_id=assignment.id)

    # Check if the student is enrolled in this class
    try:
        student = Student.objects.get(user=request.user)

        # Verify student is enrolled in the class
        if not student.enrolled_subjects.filter(id=assignment.class_subject.id).exists():
            messages.error(request, "You are not enrolled in this class.")
            return redirect('assignments:assignment_list')

        # Check if student has already submitted - required for retake
        existing_submission = StudentSubmission.objects.filter(
            student=student,
            assignment=assignment
        ).first()

        # If no existing submission, redirect to regular take_quiz
        if not existing_submission:
            messages.info(request, "You haven't taken this quiz yet. Redirecting to take quiz.")
            return redirect('assignments:take_quiz', assignment_id=assignment.id)

        # Check if quiz is still active and not past due date
        now = timezone.now()
        time_remaining = assignment.due_date - now

        if time_remaining.days < 0:
            messages.error(request, "The deadline for this quiz has passed. You cannot retake it.")
            return redirect('assignments:assignment_detail', assignment_id=assignment.id)

        # Check if the student can attempt the quiz using QuizAttempt model
        can_attempt, message = QuizAttempt.can_attempt_quiz(student, assignment)

        if not can_attempt:
            messages.error(request, message)
            return redirect('assignments:assignment_detail', assignment_id=assignment.id)

        # Get attempt count from QuizAttempt model
        attempt_count = QuizAttempt.get_student_attempts(student, assignment) + 1

        # Create a new attempt record
        quiz_attempt = QuizAttempt.create_attempt(student, assignment)

        # Get all questions
        all_questions = list(Question.objects.filter(assignment=assignment).order_by('order'))

        # Calculate total points for the assignment's max_score
        total_quiz_points = sum(q.points for q in all_questions)

        # Implement question bank functionality - if questions_to_display is set and less than total questions
        if assignment.questions_to_display and len(all_questions) > assignment.questions_to_display:
            import random
            # Determine which questions to show
            if assignment.randomize_questions:
                # Randomly select the specified number of questions
                questions = random.sample(all_questions, assignment.questions_to_display)
                # Sort selected questions by their original order
                questions.sort(key=lambda q: q.order)
            else:
                # If randomization is disabled but question bank is enabled,
                # take the first N questions in original order
                questions = all_questions[:assignment.questions_to_display]

            # Calculate adjusted max_score based on selected questions
            adjusted_max_score = sum(q.points for q in questions)
        else:
            questions = all_questions
            adjusted_max_score = total_quiz_points

        # Process new quiz submission
        if request.method == 'POST':
            # Delete previous answers to allow for new submission
            # We keep the original submission record for tracking attempts
            StudentAnswer.objects.filter(submission=existing_submission).delete()

            # Update submission date
            existing_submission.submission_date = now
            existing_submission.is_graded = False
            existing_submission.score = None
            existing_submission.save()

            # Process answers for each question
            # For each MCQ question, shuffle its choices only if randomize_choices is enabled
            for question in questions:
                if question.question_type == 'MCQ':
                    choices = list(Choice.objects.filter(question=question))
                    if assignment.randomize_choices:
                        # Shuffle the choices
                        import random
                        random.shuffle(choices)
                    # Attach choices to the question (shuffled or original order)
                    question.shuffled_choices = choices

                answer = None

                # Handle different question types
                if question.question_type == 'MCQ':
                    # Multiple choice
                    choice_id = request.POST.get(f'choice_{question.id}')
                    if choice_id:
                        try:
                            selected_choice = Choice.objects.get(id=choice_id, question=question)
                            answer = StudentAnswer.objects.create(
                                submission=existing_submission,
                                question=question,
                                selected_choice=selected_choice
                            )
                        except Choice.DoesNotExist:
                            # Invalid choice, create empty answer
                            answer = StudentAnswer.objects.create(
                                submission=existing_submission,
                                question=question
                            )
                    else:
                        # No choice selected, create empty answer
                        answer = StudentAnswer.objects.create(
                            submission=existing_submission,
                            question=question
                        )

                elif question.question_type in ['SHORT', 'LONG']:
                    # Text answers
                    text_answer = request.POST.get(f'text_{question.id}', '').strip()
                    answer = StudentAnswer.objects.create(
                        submission=existing_submission,
                        question=question,
                        text_answer=text_answer
                    )

                elif question.question_type == 'FILE':
                    # File upload
                    file_field = f'file_{question.id}'
                    if file_field in request.FILES:
                        answer = StudentAnswer.objects.create(
                            submission=existing_submission,
                            question=question,
                            file_answer=request.FILES[file_field]
                        )
                    else:
                        # No file uploaded, create empty answer
                        answer = StudentAnswer.objects.create(
                            submission=existing_submission,
                            question=question
                        )

            # Auto-grade MCQ questions
            mcq_answers = StudentAnswer.objects.filter(
                submission=existing_submission,
                question__question_type='MCQ'
            )

            # Track auto-grading stats
            auto_graded_count = 0
            auto_correct_count = 0
            total_points = 0

            # Auto-grade each MCQ answer
            for answer in mcq_answers:
                # Get the correct choice for this MCQ question
                try:
                    correct_choice = Choice.objects.get(question=answer.question, is_correct=True)

                    # Compare with student's answer
                    if answer.selected_choice and answer.selected_choice.id == correct_choice.id:
                        # Correct answer
                        answer.is_correct = True
                        answer.score = answer.question.points
                        total_points += float(answer.question.points)
                        auto_correct_count += 1

                        if answer.question.show_feedback:
                            answer.feedback = "Correct answer!"
                    else:
                        # Incorrect answer
                        answer.is_correct = False
                        answer.score = 0

                        if answer.question.show_feedback:
                            answer.feedback = f"Incorrect. The correct answer is: {correct_choice.choice_text}"

                    answer.save()
                    auto_graded_count += 1

                except Choice.DoesNotExist:
                    # Skip questions with no correct answer defined
                    continue

            # If we have auto-graded some questions, update the submission
            if auto_graded_count > 0:
                # Check if all questions were MCQ and auto-graded
                total_questions = StudentAnswer.objects.filter(submission=existing_submission).count()

                if auto_graded_count == total_questions:
                    # All questions were auto-graded, mark the submission as graded
                    existing_submission.is_graded = True
                    existing_submission.score = total_points
                    existing_submission.teacher_comment = f"Auto-graded quiz (retake #{attempt_count}). Score: {auto_correct_count}/{auto_graded_count} correct."
                    existing_submission.save()

            # Mark the quiz attempt as completed
            quiz_attempt.complete_attempt()

            messages.success(request, f"Quiz retake #{attempt_count} submitted successfully!")
            if auto_graded_count > 0:
                messages.info(request, f"{auto_graded_count} multiple-choice questions were automatically graded.")

            # Redirect back to take_quiz with retake flag
            return redirect('assignments:take_quiz', assignment_id=assignment.id)

        # Extract numeric grade if student.grade is a ClassRoom object
        numeric_grade = None
        if student.grade:
            if hasattr(student.grade, 'grade'):
                try:
                    # Strip non-numeric characters and convert to int
                    import re
                    grade_str = student.grade.grade
                    numeric_match = re.search(r'\d+', grade_str)
                    if numeric_match:
                        numeric_grade = int(numeric_match.group())
                except (ValueError, AttributeError):
                    # If conversion fails, leave as None
                    pass

        # Get maximum attempts based on grade level
        max_attempts = 3  # Default for elementary
        if numeric_grade and numeric_grade >= 9:  # High school
            max_attempts = 2
        elif numeric_grade and numeric_grade >= 6:  # Middle school
            max_attempts = 2

        # Prepare context for template with retake information
        context = {
            'assignment': assignment,
            'questions': questions,
            'existing_submission': None,  # Set to None so the form displays for retake
            'time_remaining': time_remaining,
            'is_retake': True,
            'attempt_number': attempt_count,
            'max_attempts': max_attempts,
        }

        return render(request, 'assignments/take_quiz.html', context)

    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('dashboard:index')

@login_required
@user_passes_test(lambda u: is_student(u))
def quiz_result(request, assignment_id):
    """
    Display quiz results to students after submission.
    Shows correct/incorrect answers, score, and feedback.
    """
    # Get the assignment
    assignment = get_object_or_404(Assignment, id=assignment_id)

    # Verify this is a quiz assignment
    if assignment.assignment_type != 'QUIZ':
        messages.error(request, "This is not a quiz assignment.")
        return redirect('assignments:assignment_detail', assignment_id=assignment.id)

    # Check if the student is enrolled in this class and has a submission
    try:
        student = Student.objects.get(user=request.user)

        # Verify student is enrolled in the class
        if not student.enrolled_subjects.filter(id=assignment.class_subject.id).exists():
            messages.error(request, "You are not enrolled in this class.")
            return redirect('assignments:assignment_list')

        # Get the student's submission
        submission = StudentSubmission.objects.filter(
            student=student,
            assignment=assignment
        ).first()

        if not submission:
            messages.error(request, "You haven't submitted this quiz yet.")
            return redirect('assignments:assignment_detail', assignment_id=assignment.id)

        # Get all answers for this submission
        answers = StudentAnswer.objects.filter(
            submission=submission
        ).order_by('question__order')

        # Calculate statistics
        answered_count = 0
        correct_count = 0
        incorrect_count = 0

        for answer in answers:
            # Count answered questions
            if answer.selected_choice or answer.text_answer or answer.file_answer:
                answered_count += 1

            # Count correct/incorrect answers if graded
            if submission.is_graded:
                if answer.is_correct:
                    correct_count += 1
                else:
                    incorrect_count += 1

        # Prepare context
        context = {
            'assignment': assignment,
            'submission': submission,
            'answers': answers,
            'answered_count': answered_count,
            'correct_count': correct_count,
            'incorrect_count': incorrect_count
        }

        return render(request, 'assignments/quiz_result.html', context)

    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('dashboard:index')

@login_required
def view_quiz_result(request, quiz_id):
    return render(request, 'assignments/view_quiz_result.html')

# Report card views
@login_required
def report_card_list(request):
    """
    Display list of report cards with filtering and performance metrics
    """
    user = request.user

    # Check if user is a teacher - teachers should use generate_report_cards instead
    if is_teacher(user):
        messages.warning(request, "Teachers cannot access this page directly. Please use the 'Generate Report Cards' feature to create and view report cards.")
        return redirect('assignments:generate_report_cards')

    # Get filter parameters
    time_period = request.GET.get('time_period', 'all')
    student_id = request.GET.get('student_id')
    class_id = request.GET.get('class_id')

    # Base query
    report_cards = ReportCard.objects.all()

    # Filter based on user role
    if is_student(user):
        report_cards = report_cards.filter(student__user=user)
    elif is_parent(user):
        parent = Parent.objects.get(user=user)
        report_cards = report_cards.filter(student__in=parent.children.all())

    # Apply time period filter
    today = timezone.now().date()
    if time_period == 'weekly':
        start_date = today - timedelta(days=7)
        report_cards = report_cards.filter(created_at__gte=start_date)
    elif time_period == 'monthly':
        start_date = today.replace(day=1)
        report_cards = report_cards.filter(created_at__gte=start_date)
    elif time_period == 'termly':
        # Assuming terms are Jan-Apr, May-Aug, Sep-Dec
        month = today.month
        if month <= 4:
            start_date = today.replace(month=1, day=1)
            end_date = today.replace(month=4, day=30)
        elif month <= 8:
            start_date = today.replace(month=5, day=1)
            end_date = today.replace(month=8, day=31)
        else:
            start_date = today.replace(month=9, day=1)
            end_date = today.replace(month=12, day=31)
        report_cards = report_cards.filter(created_at__range=[start_date, end_date])
    elif time_period == 'yearly':
        start_date = today.replace(month=1, day=1)
        end_date = today.replace(month=12, day=31)
        report_cards = report_cards.filter(created_at__range=[start_date, end_date])

    # Apply student filter
    if student_id:
        report_cards = report_cards.filter(student_id=student_id)

    # Apply class filter
    if class_id:
        report_cards = report_cards.filter(student__classroom_id=class_id)

    # Calculate statistics
    total_report_cards = report_cards.count()
    average_score = report_cards.aggregate(avg_score=Avg('average_score'))['avg_score'] or 0
    highest_score = report_cards.aggregate(max_score=Max('average_score'))['max_score'] or 0
    lowest_score = report_cards.aggregate(min_score=Min('average_score'))['min_score'] or 0

    # Get available students and classes for filters
    if is_admin(user):
        students = Student.objects.all().order_by('user__first_name')
        classes = ClassRoom.objects.all().order_by('name')
    elif is_teacher(user):
        teacher = user.teacher
        students = Student.objects.filter(
            Q(classroom__class_teacher=teacher) |
            Q(classroom__subjects__teacher=teacher)
        ).distinct().order_by('user__first_name')
        classes = ClassRoom.objects.filter(
            Q(class_teacher=teacher) |
            Q(subjects__teacher=teacher)
        ).distinct().order_by('name')
    elif is_parent(user):
        parent = Parent.objects.get(user=user)
        students = parent.children.all().order_by('user__first_name')
        classes = ClassRoom.objects.filter(
            subjects__students__in=students
        ).distinct().order_by('name')
    else:
        students = Student.objects.none()
        classes = ClassRoom.objects.none()

    context = {
        'report_cards': report_cards,
        'total_report_cards': total_report_cards,
        'average_score': round(average_score, 2),
        'highest_score': highest_score,
        'lowest_score': lowest_score,
        'time_period': time_period,
        'student_id': student_id,
        'class_id': class_id,
        'students': students,
        'classes': classes,
    }

    return render(request, 'assignments/report_card_list.html', context)

@login_required
@user_passes_test(lambda u: is_admin(u))
def generate_report_cards(request):
    """
    Handle report card generation with class and term selection.
    Shows student grades and allows generating report cards.
    Only administrators can generate report cards.
    """
    # If a teacher somehow tries to access this page, show an access denied message
    if is_teacher(request.user):
        messages.error(request, "Access denied. Only administrators can generate report cards.")
        return redirect('assignments:report_card_list')

    # Get classrooms based on user role
    if is_admin(request.user):
        classrooms = ClassRoom.objects.all().order_by('name', 'section')
    elif is_teacher(request.user):
        teacher = request.user.teacher
        # Only show classrooms where this teacher is the class teacher
        classrooms = ClassRoom.objects.filter(class_teacher=teacher).order_by('name', 'section')

    # Get academic years and terms from settings
    from ricas_school_manager.settings import ACADEMIC_YEARS, ACADEMIC_TERMS
    academic_years = ACADEMIC_YEARS
    academic_terms = ACADEMIC_TERMS

    # Get assessment weight configurations
    weight_configurations = AssessmentWeightConfiguration.objects.all()

    # Process form submission
    if request.method == 'POST':
        classroom_id = request.POST.get('classroom')
        academic_term = request.POST.get('academic_term')
        academic_year = request.POST.get('academic_year')

        # Validate required fields
        if not (classroom_id and academic_term and academic_year):
            messages.error(request, "Please select classroom, term, and academic year.")
            return render(request, 'assignments/generate_report_cards.html', {
                'classrooms': classrooms,
                'academic_years': academic_years,
                'academic_terms': academic_terms
            })

        # Get selected classroom
        try:
            classroom = ClassRoom.objects.get(id=classroom_id)
        except ClassRoom.DoesNotExist:
            messages.error(request, "Selected classroom does not exist.")
            return render(request, 'assignments/generate_report_cards.html', {
                'classrooms': classrooms,
                'academic_years': academic_years,
                'academic_terms': academic_terms
            })

        # Check if teacher has permission to generate report cards for this class
        if is_teacher(request.user):
            teacher = request.user.teacher
            if classroom.class_teacher != teacher:
                messages.error(request, "You can only generate report cards for classes where you are the class teacher.")
                return redirect('assignments:generate_report_cards')

        # Get students in the class
        students = Student.objects.filter(classroom=classroom)

        # Check if students were selected or all students should be included
        student_ids = request.POST.getlist('students')
        regenerate = 'regenerate' in request.POST
        include_teacher_comments = 'include_teacher_comments' in request.POST
        notify_parents = 'notify_parents' in request.POST

        if not students.exists():
            messages.error(request, "No students found in the selected class.")
            return render(request, 'assignments/generate_report_cards.html', {
                'classrooms': classrooms,
                'academic_years': academic_years,
                'academic_terms': academic_terms
            })

        # Filter students if specific ones were selected
        if student_ids:
            students = students.filter(id__in=student_ids)

        # Track report card generation stats
        report_cards_created = 0
        report_cards_updated = 0
        notification_count = 0

        # Bulk data processing
        student_data = {}

        # Get all class subjects for this classroom
        subjects = ClassSubject.objects.filter(classroom=classroom)

        # Get existing report cards for this term/year to avoid duplicates
        existing_report_cards = ReportCard.objects.filter(
            student__classroom=classroom,
            academic_term=academic_term,
            academic_year=academic_year
        )
        existing_map = {rc.student_id: rc for rc in existing_report_cards}

        # Prepare attendance data for all students in the class
        from django.db.models import Avg, Sum, Count
        from attendance.models import AttendanceRecord, StudentAttendance

        # Calculate attendance for each student
        attendance_data = {}
        for student in students:
            # Get date range for the academic term
            term_dates = get_term_date_range(academic_term, academic_year)

            if term_dates:
                start_date, end_date = term_dates

                # Get attendance records in date range
                attendance_records = AttendanceRecord.objects.filter(
                    date__range=[start_date, end_date],
                    classroom=classroom
                )

                # Get attendance for this student
                student_attendance = StudentAttendance.objects.filter(
                    record__in=attendance_records,
                    student=student
                )

                school_days = attendance_records.count()
                present_days = student_attendance.filter(status='present').count()
                absent_days = student_attendance.filter(status='absent').count()
                late_days = student_attendance.filter(status='late').count()

                attendance_data[student.id] = {
                    'school_days': school_days,
                    'present_days': present_days,
                    'absent_days': absent_days,
                    'late_days': late_days
                }

        # Get all grades for students in the selected class
        from django.db.models import Q
        grades = Grade.objects.filter(
            Q(student__in=students) &
            (Q(term=academic_term) | Q(term__isnull=True)) &
            (Q(academic_year=academic_year) | Q(academic_year__isnull=True))
        ).select_related('class_subject', 'student')

        # Group grades by student and subject
        grade_data = {}
        for grade in grades:
            student_id = grade.student.id
            if student_id not in grade_data:
                grade_data[student_id] = {}

            # Use subject name as key
            subject_name = grade.class_subject.subject.name
            if subject_name not in grade_data[student_id]:
                grade_data[student_id][subject_name] = []

            grade_data[student_id][subject_name].append(grade)

        # Generate/update report cards for each student
        for student in students:
            # Check if report card already exists
            if student.id in existing_map and not regenerate:
                # Skip this student if regenerate is False
                continue

            report_card = existing_map.get(student.id)
            action = "updated"

            if not report_card:
                # Create new report card
                report_card = ReportCard(
                    student=student,
                    academic_term=academic_term,
                    academic_year=academic_year
                )
                action = "created"
                report_cards_created += 1
            else:
                report_cards_updated += 1

            # Update attendance data
            if student.id in attendance_data:
                attendance = attendance_data[student.id]
                report_card.school_days = attendance['school_days']
                report_card.present_days = attendance['present_days']
                report_card.absent_days = attendance['absent_days']
                report_card.late_days = attendance['late_days']

            # Get teacher comments if option is selected
            if include_teacher_comments:
                teacher_comments = request.POST.get(f'teacher_comments_{student.id}', '')
                principal_comments = request.POST.get(f'principal_comments_{student.id}', '')

                report_card.teacher_comments = teacher_comments
                report_card.principal_comments = principal_comments

            # Process academic data for each subject
            total_score = 0
            subjects_count = 0

            # Get student's grades
            if student.id in grade_data:
                student_grades = grade_data[student.id]

                for subject_name, grades in student_grades.items():
                    # Calculate CA score (average of all non-exam grades)
                    ca_grades = [g for g in grades if g.grade_type != 'examination']
                    exam_grades = [g for g in grades if g.grade_type == 'examination']

                    ca_score = 0
                    if ca_grades:
                        ca_score = sum(g.score for g in ca_grades) / len(ca_grades)

                    exam_score = 0
                    if exam_grades:
                        exam_score = sum(g.score for g in exam_grades) / len(exam_grades)

                    # Calculate total score for this subject
                    subject_total = ca_score + exam_score

                    # Add to report card data
                    report_card.set_subject_grade(
                        subject=subject_name,
                        ca_score=ca_score,
                        exam_score=exam_score,
                        total=subject_total
                    )

                    # Add to total for average calculation
                    total_score += subject_total
                    subjects_count += 1

            # Calculate average score
            if subjects_count > 0:
                report_card.average_score = round(total_score / subjects_count, 2)

            # Save the report card
            report_card.save()

            # Send notification to parents if option is selected
            if notify_parents:
                try:
                    # Import notification utility
                    from communications.utils import create_notification

                    # Get parent(s) of this student
                    parents = Parent.objects.filter(children=student)

                    for parent in parents:
                        # Create notification for each parent
                        notification_message = f"Report card for {student.user.get_full_name()} is now available for {academic_term} {academic_year}."
                        create_notification(
                            user=parent.user,
                            message=notification_message,
                            link=f"/assignments/report_card/{report_card.id}/",
                            sender=request.user
                        )
                        notification_count += 1
                except Exception as e:
                    # Log error but continue processing
                    print(f"Error sending notification for student {student.id}: {str(e)}")

        # Show success message with stats
        if report_cards_created > 0 or report_cards_updated > 0:
            message_parts = []
            if report_cards_created > 0:
                message_parts.append(f"{report_cards_created} report cards created")
            if report_cards_updated > 0:
                message_parts.append(f"{report_cards_updated} report cards updated")
            if notification_count > 0:
                message_parts.append(f"{notification_count} notifications sent")

            messages.success(request, f"Report cards processed successfully: {', '.join(message_parts)}.")

            # Redirect to report card list
            return redirect('assignments:report_card_list')
        else:
            messages.info(request, "No report cards were created or updated. They may already exist for the selected term.")

    context = {
        'classrooms': classrooms,
        'academic_years': academic_years,
        'academic_terms': academic_terms,
    }

    return render(request, 'assignments/generate_report_cards.html', context)

def get_term_date_range(term, year):
    """
    Helper function to calculate date range for a given academic term and year.
    Returns (start_date, end_date) tuple.
    """
    from datetime import datetime, date

    # Parse the year from the string representation
    year_number = int(year.split('-')[0])

    # Define term date ranges
    if term == 'First Term':
        # First term is typically Sept-Dec of the previous year
        start_date = date(year_number, 9, 1)
        end_date = date(year_number, 12, 31)
    elif term == 'Second Term':
        # Second term is typically Jan-April of the current year
        start_date = date(year_number + 1, 1, 1)
        end_date = date(year_number + 1, 4, 30)
    elif term == 'Third Term':
        # Third term is typically May-August of the current year
        start_date = date(year_number + 1, 5, 1)
        end_date = date(year_number + 1, 8, 31)
    else:
        # For custom terms or if term is not recognized
        return None

    return (start_date, end_date)

@login_required
@user_passes_test(lambda u: is_admin(u))
def generate_report_card(request, student_id):
    """
    Create or update a report card for an individual student.
    Includes academic performance, attendance, behavior assessment, and teacher comments.
    Only administrators can generate report cards.
    """
    # Get the student
    student = get_object_or_404(Student, id=student_id)

    # If a teacher somehow tries to access this page, show an access denied message
    if is_teacher(request.user):
        messages.error(request, "Access denied. Only administrators can generate report cards.")
        return redirect('assignments:report_card_list')

    # Get available academic terms and years
    from ricas_school_manager.settings import ACADEMIC_TERMS, ACADEMIC_YEARS

    # Get the student's subjects
    subjects = ClassSubject.objects.filter(students=student)

    # Process form submission
    if request.method == 'POST':
        # Get form data
        academic_term = request.POST.get('academic_term')
        academic_year = request.POST.get('academic_year')
        teacher_comments = request.POST.get('teacher_comments')
        principal_comments = request.POST.get('principal_comments')
        promoted_to_next_class = 'promoted_to_next_class' in request.POST

        # Validate required fields
        if not (academic_term and academic_year):
            messages.error(request, "Please select academic term and year.")
            return render(request, 'assignments/generate_report_card.html', {
                'student': student,
                'subjects': subjects,
                'academic_terms': ACADEMIC_TERMS,
                'academic_years': ACADEMIC_YEARS
            })

        # Check if a report card already exists for this student, term and year
        report_card = ReportCard.objects.filter(
            student=student,
            academic_term=academic_term,
            academic_year=academic_year
        ).first()

        # Create new or update existing
        if report_card:
            action = "updated"
        else:
            report_card = ReportCard(
                student=student,
                academic_term=academic_term,
                academic_year=academic_year
            )
            action = "created"

        # Update common fields
        report_card.teacher_comments = teacher_comments
        report_card.principal_comments = principal_comments
        report_card.promoted_to_next_class = promoted_to_next_class

        # Process academic data for each subject
        total_score = 0
        subjects_count = 0

        for subject in subjects:
            subject_id = subject.id

            # Get subject scores from form
            ca_score_key = f'ca_score_{subject_id}'
            exam_score_key = f'exam_score_{subject_id}'

            if ca_score_key in request.POST and exam_score_key in request.POST:
                try:
                    ca_score = float(request.POST.get(ca_score_key, 0))
                    exam_score = float(request.POST.get(exam_score_key, 0))

                    # Calculate total score for this subject
                    subject_total = ca_score + exam_score

                    # Add to report card data
                    report_card.set_subject_grade(
                        subject=subject.subject.name,
                        ca_score=ca_score,
                        exam_score=exam_score,
                        total=subject_total
                    )

                    # Add to total for average calculation
                    total_score += subject_total
                    subjects_count += 1
                except ValueError:
                    # Skip invalid scores
                    pass

        # Calculate average score
        if subjects_count > 0:
            report_card.average_score = round(total_score / subjects_count, 2)

        # Process attendance data
        if 'calculate_attendance' in request.POST:
            # Auto-calculate attendance based on records
            report_card.calculate_attendance()
        else:
            # Use manual attendance values
            try:
                report_card.school_days = int(request.POST.get('school_days', 0))
                report_card.present_days = int(request.POST.get('present_days', 0))
                report_card.absent_days = int(request.POST.get('absent_days', 0))
                report_card.late_days = int(request.POST.get('late_days', 0))
            except ValueError:
                # Use defaults if invalid
                report_card.school_days = 0
                report_card.present_days = 0
                report_card.absent_days = 0
                report_card.late_days = 0

        # Process behavior/character assessment
        report_card.punctuality_rating = request.POST.get('punctuality_rating', 1)
        report_card.neatness_rating = request.POST.get('neatness_rating', 1)
        report_card.honesty_rating = request.POST.get('honesty_rating', 1)
        report_card.relationship_with_peers_rating = request.POST.get('relationship_with_peers_rating', 1)
        report_card.leadership_rating = request.POST.get('leadership_rating', 1)
        report_card.responsibility_rating = request.POST.get('responsibility_rating', 1)
        report_card.conduct_rating = request.POST.get('conduct_rating', 1)
        report_card.attitude_rating = request.POST.get('attitude_rating', 1)
        report_card.participation_rating = request.POST.get('participation_rating', 1)
        report_card.organization_rating = request.POST.get('organization_rating', 1)
        report_card.initiative_rating = request.POST.get('initiative_rating', 1)
        report_card.verbal_skills_rating = request.POST.get('verbal_skills_rating', 1)
        report_card.skills_comments = request.POST.get('skills_comments', '')

        # Save the report card
        report_card.save()

        # Handle notification
        if 'notify_parent' in request.POST:
            try:
                # Import notification utility
                from communications.utils import create_notification

                # Get parent(s) of this student
                parents = Parent.objects.filter(children=student)

                for parent in parents:
                    # Create notification for each parent
                    notification_message = f"Report card for {student.user.get_full_name()} is now available for {academic_term} {academic_year}."
                    create_notification(
                        user=parent.user,
                        message=notification_message,
                        link=f"/assignments/report_card/{report_card.id}/",
                        sender=request.user
                    )

                messages.success(request, f"Report card {action} successfully and parent(s) notified.")
            except Exception as e:
                messages.warning(request, f"Report card {action} successfully but failed to notify parent(s): {str(e)}")
        else:
            messages.success(request, f"Report card {action} successfully.")

        # Redirect to the report card detail view
        return redirect('assignments:report_card_detail', report_card_id=report_card.id)

    # GET request - show the form
    # Check if a report card already exists for the latest term
    existing_report_card = ReportCard.objects.filter(student=student).order_by('-created_at').first()

    context = {
        'student': student,
        'subjects': subjects,
        'academic_terms': ACADEMIC_TERMS,
        'academic_years': ACADEMIC_YEARS,
        'existing_report_card': existing_report_card
    }

    return render(request, 'assignments/generate_report_card.html', context)

@login_required
def report_card_detail(request, report_card_id):
    """
    Display detailed view of a student's report card
    """
    try:
        report_card = ReportCard.objects.get(id=report_card_id)

        # Check permissions
        if not (is_admin(request.user) or
                is_teacher(request.user) or
                (is_student(request.user) and report_card.student.user == request.user) or
                (is_parent(request.user) and report_card.student in Parent.objects.get(user=request.user).children.all())):
            messages.error(request, "You don't have permission to view this report card.")
            return redirect('assignments:report_card_list')

        context = {
            'report_card': report_card,
            'title': f'Report Card - {report_card.student.user.get_full_name()}',
        }

        return render(request, 'assignments/report_card_detail.html', context)

    except ReportCard.DoesNotExist:
        messages.error(request, "Report card not found.")
        return redirect('assignments:report_card_list')

@login_required
def view_report_card(request, report_card_id):
    """
    Display a detailed view of a student's report card
    """
    try:
        report_card = ReportCard.objects.get(id=report_card_id)

        # Check permissions
        if not (is_admin(request.user) or
                is_teacher(request.user) or
                (is_student(request.user) and report_card.student.user == request.user) or
                (is_parent(request.user) and report_card.student in Parent.objects.get(user=request.user).children.all())):
            messages.error(request, "You don't have permission to view this report card.")
            return redirect('assignments:report_card_list')

        context = {
            'report_card': report_card,
            'title': f'Report Card - {report_card.student.user.get_full_name()}',
        }

        return render(request, 'assignments/view_report_card.html', context)

    except ReportCard.DoesNotExist:
        messages.error(request, "Report card not found.")
        return redirect('assignments:report_card_list')

@login_required
def print_report_card(request, report_card_id):
    """
    Display a print-friendly version of the report card.
    """
    try:
        report_card = ReportCard.objects.get(id=report_card_id)

        # Check permissions
        if not (is_admin(request.user) or
                is_teacher(request.user) or
                (is_student(request.user) and report_card.student.user == request.user) or
                (is_parent(request.user) and report_card.student in Parent.objects.get(user=request.user).children.all())):
            messages.error(request, "You don't have permission to view this report card.")
            return redirect('assignments:report_card_list')

        context = {
            'report_card': report_card,
            'title': f'Print Report Card - {report_card.student.user.get_full_name()}',
        }

        return render(request, 'assignments/print_report_card.html', context)

    except ReportCard.DoesNotExist:
        messages.error(request, "Report card not found.")
        return redirect('assignments:report_card_list')

@login_required
def export_report_card(request, report_card_id, format_type):
    """
    Export a report card as PDF or Excel.

    Args:
        request: The HTTP request
        report_card_id: The ID of the report card to export
        format_type: The format to export to ('pdf' or 'excel')
    """
    try:
        report_card = ReportCard.objects.get(id=report_card_id)

        # Check permissions
        if not (is_admin(request.user) or
                is_teacher(request.user) or
                (is_student(request.user) and report_card.student.user == request.user) or
                (is_parent(request.user) and report_card.student in Parent.objects.get(user=request.user).children.all())):
            messages.error(request, "You don't have permission to export this report card.")
            return redirect('assignments:report_card_list')

        student_name = report_card.student.user.get_full_name()
        term_info = f"{report_card.academic_term}_{report_card.academic_year}"
        filename = f"Report_Card_{student_name}_{term_info}".replace(" ", "_")

        if format_type == 'pdf':
            # Create PDF export using WeasyPrint
            from django.http import HttpResponse
            from django.template.loader import get_template
            from weasyprint import HTML, CSS
            from django.conf import settings
            import tempfile

            # Render the template with the report card data
            template = get_template('assignments/print_report_card.html')
            context = {
                'report_card': report_card,
                'title': f'Report Card - {student_name}',
                'export_mode': True
            }
            html_string = template.render(context)

            # Create a temporary file to store the PDF
            with tempfile.NamedTemporaryFile(delete=True) as output:
                # Generate PDF using WeasyPrint
                html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
                css = CSS(string='@page { size: A4; margin: 1cm; }')
                html.write_pdf(output.name, stylesheets=[css])

                # Read the PDF and create response
                with open(output.name, 'rb') as pdf_file:
                    response = HttpResponse(pdf_file.read(), content_type='application/pdf')
                    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
                    return response

        elif format_type == 'excel':
            # Create Excel export using openpyxl
            from django.http import HttpResponse
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter

            # Create a new workbook and select the active sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Report Card"

            # Set column widths
            for col_idx in range(1, 15):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15

            # Header styles
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            centered = Alignment(horizontal='center', vertical='center')

            # Add school and report card info
            ws.merge_cells('A1:E1')
            ws['A1'] = 'STUDENT REPORT CARD'
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = centered

            # Student Information
            row = 3
            ws['A' + str(row)] = 'Student Name:'
            ws['B' + str(row)] = student_name
            ws['A' + str(row)].font = header_font

            row += 1
            ws['A' + str(row)] = 'Student ID:'
            ws['B' + str(row)] = report_card.student.student_id
            ws['A' + str(row)].font = header_font

            row += 1
            ws['A' + str(row)] = 'Class:'
            ws['B' + str(row)] = str(report_card.student.classroom)
            ws['A' + str(row)].font = header_font

            row += 1
            ws['A' + str(row)] = 'Term:'
            ws['B' + str(row)] = report_card.academic_term
            ws['A' + str(row)].font = header_font

            row += 1
            ws['A' + str(row)] = 'Academic Year:'
            ws['B' + str(row)] = report_card.academic_year
            ws['A' + str(row)].font = header_font

            # Attendance Section
            row += 2
            ws.merge_cells(f'A{row}:E{row}')
            ws['A' + str(row)] = 'ATTENDANCE SUMMARY'
            ws['A' + str(row)].font = header_font
            ws['A' + str(row)].alignment = centered
            ws['A' + str(row)].fill = header_fill

            row += 1
            headers = ['School Days', 'Present', 'Absent', 'Late', 'Attendance %']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.alignment = centered

            row += 1
            attendance_data = [
                report_card.school_days,
                report_card.present_days,
                report_card.absent_days,
                report_card.late_days,
                f"{report_card.attendance_percentage:.1f}%"
            ]
            for col_idx, value in enumerate(attendance_data, 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = value
                cell.alignment = centered

            # Academic Performance
            row += 2
            ws.merge_cells(f'A{row}:E{row}')
            ws['A' + str(row)] = 'ACADEMIC PERFORMANCE'
            ws['A' + str(row)].font = header_font
            ws['A' + str(row)].alignment = centered
            ws['A' + str(row)].fill = header_fill

            row += 1
            headers = ['Subject', 'CA Score', 'Exam Score', 'Total', 'Grade']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.alignment = centered

            # Get subject grades
            subjects_data = report_card.get_subject_grades()

            for subject_data in subjects_data:
                row += 1
                data = [
                    subject_data['subject'],
                    subject_data['ca_score'],
                    subject_data['exam_score'],
                    subject_data['total'],
                    subject_data.get('grade', '')
                ]
                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = value
                    cell.alignment = centered

            # Add average score
            row += 1
            ws.cell(row=row, column=1).value = "Average Score:"
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=4).value = report_card.average_score
            ws.cell(row=row, column=4).alignment = centered

            # Skills Assessment
            row += 2
            ws.merge_cells(f'A{row}:E{row}')
            ws['A' + str(row)] = 'SKILLS & CONDUCT ASSESSMENT'
            ws['A' + str(row)].font = header_font
            ws['A' + str(row)].alignment = centered
            ws['A' + str(row)].fill = header_fill

            row += 1
            headers = ['Skill', 'Rating (1-5)']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.alignment = centered

            # Skills data
            skills_data = [
                ('Punctuality', report_card.punctuality_rating),
                ('Neatness', report_card.neatness_rating),
                ('Honesty', report_card.honesty_rating),
                ('Relationship with Peers', report_card.relationship_with_peers_rating),
                ('Leadership', report_card.leadership_rating),
                ('Responsibility', report_card.responsibility_rating),
                ('Conduct', report_card.conduct_rating),
                ('Attitude', report_card.attitude_rating),
                ('Participation', report_card.participation_rating),
                ('Organization', report_card.organization_rating),
                ('Initiative', report_card.initiative_rating),
                ('Verbal Skills', report_card.verbal_skills_rating)
            ]

            for skill, rating in skills_data:
                row += 1
                ws.cell(row=row, column=1).value = skill
                ws.cell(row=row, column=2).value = rating
                ws.cell(row=row, column=2).alignment = centered

            # Comments Section
            row += 2
            ws.merge_cells(f'A{row}:E{row}')
            ws['A' + str(row)] = 'COMMENTS'
            ws['A' + str(row)].font = header_font
            ws['A' + str(row)].alignment = centered
            ws['A' + str(row)].fill = header_fill

            row += 1
            ws.cell(row=row, column=1).value = "Teacher's Comments:"
            ws.cell(row=row, column=1).font = header_font

            row += 1
            ws.merge_cells(f'A{row}:E{row}')
            ws.cell(row=row, column=1).value = report_card.teacher_comments
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 60

            row += 2
            ws.cell(row=row, column=1).value = "Principal's Comments:"
            ws.cell(row=row, column=1).font = header_font

            row += 1
            ws.merge_cells(f'A{row}:E{row}')
            ws.cell(row=row, column=1).value = report_card.principal_comments
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 60

            # Promotional Status
            row += 2
            ws.cell(row=row, column=1).value = "Promotional Status:"
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=2).value = "Promoted to Next Class" if report_card.promoted_to_next_class else "Not Promoted"

            # Create response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            wb.save(response)
            return response

        else:
            messages.error(request, f"Unsupported export format: {format_type}")
            return redirect('assignments:report_card_detail', report_card_id=report_card_id)

    except ReportCard.DoesNotExist:
        messages.error(request, "Report card not found.")
        return redirect('assignments:report_card_list')
    except Exception as e:
        messages.error(request, f"Error exporting report card: {str(e)}")
        return redirect('assignments:report_card_detail', report_card_id=report_card_id)

@login_required
@user_passes_test(lambda u: is_admin(u))
def export_all_report_cards(request, format_type):
    """
    Export all report cards for a class as a single Excel workbook or ZIP of PDFs.

    Args:
        request: The HTTP request
        format_type: The format to export to ('pdf' or 'excel')
    """
    # Get filter parameters
    class_id = request.GET.get('class_id')
    academic_term = request.GET.get('academic_term')
    academic_year = request.GET.get('academic_year')

    if not (class_id and academic_term and academic_year):
        messages.error(request, "Please specify class, term and academic year for bulk export.")
        return redirect('assignments:report_card_list')

    try:
        classroom = ClassRoom.objects.get(id=class_id)

        # Get report cards for this class, term and academic year
        report_cards = ReportCard.objects.filter(
            student__classroom=classroom,
            academic_term=academic_term,
            academic_year=academic_year
        ).order_by('student__user__last_name', 'student__user__first_name')

        if not report_cards.exists():
            messages.error(request, "No report cards found for the selected class, term and academic year.")
            return redirect('assignments:report_card_list')

        filename_base = f"Report_Cards_{classroom.name}_{academic_term}_{academic_year}".replace(" ", "_")

        if format_type == 'excel':
            # Create Excel export with multiple sheets
            from django.http import HttpResponse
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter

            # Create a new workbook
            wb = openpyxl.Workbook()

            # Create a summary sheet
            summary = wb.active
            summary.title = "Summary"

            # Set column widths
            for col_idx in range(1, 10):
                summary.column_dimensions[get_column_letter(col_idx)].width = 15

            # Header styles
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            centered = Alignment(horizontal='center', vertical='center')

            # Add class info
            summary.merge_cells('A1:G1')
            summary['A1'] = f'REPORT CARD SUMMARY - {classroom.name} - {academic_term} {academic_year}'
            summary['A1'].font = Font(bold=True, size=14)
            summary['A1'].alignment = centered

            # Add headers for summary table
            row = 3
            headers = ['Student Name', 'ID', 'Average Score', 'Attendance %', 'Present Days', 'Absent Days', 'Late Days', 'Promoted']
            for col_idx, header in enumerate(headers, 1):
                cell = summary.cell(row=row, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.alignment = centered
                cell.fill = header_fill

            # Add data for each student
            for report_card in report_cards:
                row += 1
                student = report_card.student
                data = [
                    student.user.get_full_name(),
                    student.student_id,
                    report_card.average_score,
                    f"{report_card.attendance_percentage:.1f}%",
                    report_card.present_days,
                    report_card.absent_days,
                    report_card.late_days,
                    "Yes" if report_card.promoted_to_next_class else "No"
                ]
                for col_idx, value in enumerate(data, 1):
                    cell = summary.cell(row=row, column=col_idx)
                    cell.value = value
                    cell.alignment = centered

            # Calculate class average
            row += 2
            summary.cell(row=row, column=1).value = "Class Average Score:"
            summary.cell(row=row, column=1).font = header_font
            class_avg = sum(rc.average_score for rc in report_cards) / report_cards.count() if report_cards.count() > 0 else 0
            summary.cell(row=row, column=3).value = round(class_avg, 2)
            summary.cell(row=row, column=3).alignment = centered

            # Calculate attendance statistics
            row += 1
            summary.cell(row=row, column=1).value = "Class Attendance Average:"
            summary.cell(row=row, column=1).font = header_font
            attendance_avg = sum(rc.attendance_percentage for rc in report_cards) / report_cards.count() if report_cards.count() > 0 else 0
            summary.cell(row=row, column=4).value = f"{attendance_avg:.1f}%"
            summary.cell(row=row, column=4).alignment = centered

            # Calculate promotion rate
            row += 1
            summary.cell(row=row, column=1).value = "Promotion Rate:"
            summary.cell(row=row, column=1).font = header_font
            promoted_count = sum(1 for rc in report_cards if rc.promoted_to_next_class)
            promotion_rate = (promoted_count / report_cards.count()) * 100 if report_cards.count() > 0 else 0
            summary.cell(row=row, column=8).value = f"{promotion_rate:.1f}%"
            summary.cell(row=row, column=8).alignment = centered

            # Add individual student sheets
            for report_card in report_cards:
                student = report_card.student
                student_name = student.user.get_full_name()
                sheet_name = student_name[:31]  # Excel sheet names limited to 31 chars

                # Create a new sheet for this student
                ws = wb.create_sheet(title=sheet_name)

                # Set column widths
                for col_idx in range(1, 15):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15

                # Add student report card (similar to individual export)
                # Add school and report card info
                ws.merge_cells('A1:E1')
                ws['A1'] = 'STUDENT REPORT CARD'
                ws['A1'].font = Font(bold=True, size=14)
                ws['A1'].alignment = centered

                # Student Information
                row = 3
                ws['A' + str(row)] = 'Student Name:'
                ws['B' + str(row)] = student_name
                ws['A' + str(row)].font = header_font

                row += 1
                ws['A' + str(row)] = 'Student ID:'
                ws['B' + str(row)] = student.student_id
                ws['A' + str(row)].font = header_font

                row += 1
                ws['A' + str(row)] = 'Class:'
                ws['B' + str(row)] = str(student.classroom)
                ws['A' + str(row)].font = header_font

                row += 1
                ws['A' + str(row)] = 'Term:'
                ws['B' + str(row)] = report_card.academic_term
                ws['A' + str(row)].font = header_font

                row += 1
                ws['A' + str(row)] = 'Academic Year:'
                ws['B' + str(row)] = report_card.academic_year
                ws['A' + str(row)].font = header_font

                # Attendance Section
                row += 2
                ws.merge_cells(f'A{row}:E{row}')
                ws['A' + str(row)] = 'ATTENDANCE SUMMARY'
                ws['A' + str(row)].font = header_font
                ws['A' + str(row)].alignment = centered
                ws['A' + str(row)].fill = header_fill

                row += 1
                headers = ['School Days', 'Present', 'Absent', 'Late', 'Attendance %']
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = header
                    cell.font = header_font
                    cell.alignment = centered

                row += 1
                attendance_data = [
                    report_card.school_days,
                    report_card.present_days,
                    report_card.absent_days,
                    report_card.late_days,
                    f"{report_card.attendance_percentage:.1f}%"
                ]
                for col_idx, value in enumerate(attendance_data, 1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = value
                    cell.alignment = centered

                # Academic Performance
                row += 2
                ws.merge_cells(f'A{row}:E{row}')
                ws['A' + str(row)] = 'ACADEMIC PERFORMANCE'
                ws['A' + str(row)].font = header_font
                ws['A' + str(row)].alignment = centered
                ws['A' + str(row)].fill = header_fill

                row += 1
                headers = ['Subject', 'CA Score', 'Exam Score', 'Total', 'Grade']
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = header
                    cell.font = header_font
                    cell.alignment = centered

                # Get subject grades
                subjects_data = report_card.get_subject_grades()

                for subject_data in subjects_data:
                    row += 1
                    data = [
                        subject_data['subject'],
                        subject_data['ca_score'],
                        subject_data['exam_score'],
                        subject_data['total'],
                        subject_data.get('grade', '')
                    ]
                    for col_idx, value in enumerate(data, 1):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.value = value
                        cell.alignment = centered

                # Add average score
                row += 1
                ws.cell(row=row, column=1).value = "Average Score:"
                ws.cell(row=row, column=1).font = header_font
                ws.cell(row=row, column=4).value = report_card.average_score
                ws.cell(row=row, column=4).alignment = centered

                # Comments Section (simplified for bulk export)
                row += 2
                ws.merge_cells(f'A{row}:E{row}')
                ws['A' + str(row)] = 'COMMENTS'
                ws['A' + str(row)].font = header_font
                ws['A' + str(row)].alignment = centered
                ws['A' + str(row)].fill = header_fill

                row += 1
                ws.cell(row=row, column=1).value = "Teacher's Comments:"
                ws.cell(row=row, column=1).font = header_font

                row += 1
                ws.merge_cells(f'A{row}:E{row}')
                ws.cell(row=row, column=1).value = report_card.teacher_comments
                ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
                ws.row_dimensions[row].height = 40

                # Promotional Status
                row += 2
                ws.cell(row=row, column=1).value = "Promotional Status:"
                ws.cell(row=row, column=1).font = header_font
                ws.cell(row=row, column=2).value = "Promoted to Next Class" if report_card.promoted_to_next_class else "Not Promoted"

            # Create response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
            wb.save(response)
            return response

        elif format_type == 'pdf':
            # Create a ZIP file containing individual PDFs
            from django.http import HttpResponse
            from django.template.loader import get_template
            from weasyprint import HTML, CSS
            import zipfile
            import io
            import tempfile

            # Create a temporary ZIP file
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                template = get_template('assignments/print_report_card.html')

                # Create a PDF for each report card
                for report_card in report_cards:
                    student = report_card.student
                    student_name = student.user.get_full_name()
                    pdf_filename = f"Report_Card_{student_name}_{academic_term}_{academic_year}.pdf".replace(" ", "_")

                    # Render the template with the report card data
                    context = {
                        'report_card': report_card,
                        'title': f'Report Card - {student_name}',
                        'export_mode': True
                    }
                    html_string = template.render(context)

                    # Create a temporary file to store the PDF
                    with tempfile.NamedTemporaryFile(delete=True) as output:
                        # Generate PDF using WeasyPrint
                        html = HTML(string=html_string)
                        css = CSS(string='@page { size: A4; margin: 1cm; }')
                        html.write_pdf(output.name, stylesheets=[css])

                        # Read the PDF and add to ZIP
                        with open(output.name, 'rb') as pdf_file:
                            zip_file.writestr(pdf_filename, pdf_file.read())

            # Create response with the ZIP file
            zip_buffer.seek(0)
            response = HttpResponse(zip_buffer, content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{filename_base}.zip"'
            return response

        else:
            messages.error(request, f"Unsupported export format: {format_type}")
            return redirect('assignments:report_card_list')

    except ClassRoom.DoesNotExist:
        messages.error(request, "Class not found.")
        return redirect('assignments:report_card_list')
    except Exception as e:
        messages.error(request, f"Error exporting report cards: {str(e)}")
        return redirect('assignments:report_card_list')

# Gradebook views
@login_required
def gradebook(request):
    """
    Display all grades for a teacher's classes with filtering options.
    Shows grade distribution, averages, and allows drilling down into individual classes.
    Features a spreadsheet-like interface with cumulative assessment tracking.
    """
    user = request.user
    if not (is_teacher(user) or is_admin(user)):
        messages.error(request, "Only teachers and administrators can access the gradebook")
        return redirect('dashboard:index')

    # Get filter parameters
    selected_class_id = request.GET.get('class')
    selected_subject_id = request.GET.get('subject')
    selected_term = request.GET.get('term')
    selected_year = request.GET.get('year')

    # Default to current academic year if not specified
    if not selected_year:
        from datetime import datetime
        current_year = datetime.now().year
        if datetime.now().month < 9:  # Before September, use previous year's academic year
            selected_year = f"{current_year-1}-{current_year}"
        else:
            selected_year = f"{current_year}-{current_year+1}"

    # Get classes and subjects based on user role
    if is_teacher(user):
        teacher = Teacher.objects.get(user=user)
        classes = ClassRoom.objects.filter(
            Q(class_teacher=teacher) | Q(subjects__teacher=teacher)
        ).distinct()
    else:  # Admin
        classes = ClassRoom.objects.all()

    selected_class = None
    subjects = []
    grades = []
    students = []
    assessments = []
    avg_score = 0
    assessment_count = 0
    pass_rate = 0
    class_avg = 0
    grade_distribution = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}
    assessment_stats = {}

    if selected_class_id:
        try:
            selected_class = classes.get(id=selected_class_id)
            subjects = ClassSubject.objects.filter(classroom=selected_class)
            if is_teacher(user):
                subjects = subjects.filter(
                    Q(teacher=teacher) | Q(classroom__class_teacher=teacher)
                ).distinct()

            # Get students for the selected class
            students = Student.objects.filter(classroom=selected_class)

            # Get grades if subject is selected
            if selected_subject_id:
                try:
                    selected_subject = subjects.get(id=selected_subject_id)

                    # Get all assignments for this class subject
                    assignments = Assignment.objects.filter(
                        class_subject=selected_subject
                    )

                    # Apply term filter if specified
                    if selected_term:
                        # Use assignment metadata or comments to filter by term
                        filtered_assignments = []
                        for assignment in assignments:
                            if assignment.comments:
                                try:
                                    metadata = json.loads(assignment.comments)
                                    if metadata.get('term') == selected_term:
                                        filtered_assignments.append(assignment)
                                except json.JSONDecodeError:
                                    pass
                        assignments = filtered_assignments

                    # Process assignments into assessment types
                    assessment_types = {
                        'classwork': [],
                        'quiz': [],
                        'test': [],
                        'midterm': [],
                        'project': [],
                        'exam': []
                    }

                    for assignment in assignments:
                        # Determine assessment type from assignment type or metadata
                        assignment_type = assignment.assignment_type.lower()
                        assessment_type = 'classwork'  # Default

                        if assignment_type == 'homework':
                            assessment_type = 'classwork'
                        elif assignment_type == 'quiz':
                            assessment_type = 'quiz'
                        elif assignment_type == 'test':
                            assessment_type = 'test'
                        elif assignment_type == 'exam':
                            assessment_type = 'exam'

                        # Check metadata for more specific type
                        if assignment.comments:
                            try:
                                metadata = json.loads(assignment.comments)
                                if 'component' in metadata:
                                    component = metadata.get('component')
                                    if component in assessment_types:
                                        assessment_type = component
                            except json.JSONDecodeError:
                                pass

                        # Create assessment object
                        assessment = {
                            'id': assignment.id,
                            'title': assignment.title,
                            'type': assessment_type,
                            'max_points': assignment.max_score,
                            'avg_score': 0,
                            'submissions': []
                        }

                        # Get submissions for this assignment
                        submissions = StudentSubmission.objects.filter(
                            assignment=assignment,
                            student__in=students
                        )

                        # Calculate average score for this assessment
                        if submissions.exists():
                            total_score = 0
                            graded_count = 0

                            for submission in submissions:
                                if submission.is_graded:
                                    total_score += submission.score
                                    graded_count += 1

                                # Store submission data
                                assessment['submissions'].append({
                                    'student_id': submission.student.id,
                                    'score': submission.score if submission.is_graded else None
                                })

                            if graded_count > 0:
                                assessment['avg_score'] = (total_score / graded_count) / assignment.max_score * 100

                        # Add to appropriate category
                        assessment_types[assessment_type].append(assessment)

                    # Flatten assessments for the template
                    for assessment_type, assessments_list in assessment_types.items():
                        for a in assessments_list:
                            assessments.append(a)

                    # Sort by type and title
                    assessments.sort(key=lambda a: (a['type'], a['title']))

                    # Get all grades for the selected subject
                    grades = Grade.objects.filter(
                        class_subject=selected_subject,
                        student__in=students
                    )

                    if selected_term:
                        grades = grades.filter(term=selected_term)

                    # Group grades by student, type and assessment
                    student_grades = {}
                    for grade in grades:
                        if grade.student_id not in student_grades:
                            student_grades[grade.student_id] = {
                                'grades': [],
                                'by_type': {},
                                'by_assessment': {},
                                'average': 0
                            }

                        student_grades[grade.student_id]['grades'].append(grade)

                        # Extract assessment type from grade metadata
                        assessment_type = 'classwork'  # Default
                        assessment_id = None

                        if grade.submission:
                            assessment_id = grade.submission.assignment.id
                            assignment_type = grade.submission.assignment.assignment_type.lower()

                            if assignment_type == 'homework':
                                assessment_type = 'classwork'
                            elif assignment_type == 'quiz':
                                assessment_type = 'quiz'
                            elif assignment_type == 'test':
                                assessment_type = 'test'
                            elif assignment_type == 'exam':
                                assessment_type = 'exam'

                        # Add grade to the appropriate type bucket
                        if assessment_type not in student_grades[grade.student_id]['by_type']:
                            student_grades[grade.student_id]['by_type'][assessment_type] = []

                        student_grades[grade.student_id]['by_type'][assessment_type].append(grade)

                        # Add to assessment lookup
                        if assessment_id:
                            student_grades[grade.student_id]['by_assessment'][assessment_id] = grade

                    # Process student data for spreadsheet view
                    student_total_scores = 0
                    student_count_with_grades = 0
                    passing_students = 0

                    for student in students:
                        # Initialize grades object
                        student.grades = {
                            'average': 0,
                            'by_assessment': {},
                            'by_type': {}
                        }

                        if student.id in student_grades:
                            student_data = student_grades[student.id]

                            # Set assessment grades
                            for assessment in assessments:
                                if assessment['id'] in student_data['by_assessment']:
                                    grade = student_data['by_assessment'][assessment['id']]
                                    student.grades['by_assessment'][assessment['id']] = grade.score

                            # Calculate student average if they have grades
                            if student_data['grades']:
                                total_score = 0
                                total_weight = 0

                                # Calculate weighted average by type
                                type_weights = {
                                    'classwork': 0.30,
                                    'quiz': 0.15,
                                    'test': 0.15,
                                    'midterm': 0.10,
                                    'project': 0.10,
                                    'exam': 0.20
                                }

                                for assessment_type, weight in type_weights.items():
                                    if assessment_type in student_data['by_type'] and student_data['by_type'][assessment_type]:
                                        grades = student_data['by_type'][assessment_type]
                                        type_avg = sum(g.score for g in grades) / sum(g.max_score for g in grades) * 100
                                        student.grades['by_type'][assessment_type] = type_avg
                                        total_score += type_avg * weight
                                        total_weight += weight

                                if total_weight > 0:
                                    student.grades['average'] = total_score / total_weight
                                else:
                                    # Simple average if no weights
                                    grades = student_data['grades']
                                    if sum(g.max_score for g in grades) > 0:
                                        student.grades['average'] = sum(g.score for g in grades) / sum(g.max_score for g in grades) * 100

                                # Update class statistics
                                student_total_scores += student.grades['average']
                                student_count_with_grades += 1

                                # Update grade distribution
                                if student.grades['average'] >= 90:
                                    grade_distribution['A'] += 1
                                    passing_students += 1
                                elif student.grades['average'] >= 80:
                                    grade_distribution['B'] += 1
                                    passing_students += 1
                                elif student.grades['average'] >= 70:
                                    grade_distribution['C'] += 1
                                    passing_students += 1
                                elif student.grades['average'] >= 60:
                                    grade_distribution['D'] += 1
                                    passing_students += 1
                                else:
                                    grade_distribution['F'] += 1

                    # Calculate class-wide statistics
                    assessment_count = len(assessments)

                    if student_count_with_grades > 0:
                        avg_score = student_total_scores / student_count_with_grades
                        class_avg = avg_score
                        pass_rate = (passing_students / student_count_with_grades) * 100

                    # Calculate assessment type statistics
                    assessment_types_list = ['classwork', 'quiz', 'test', 'midterm', 'project', 'exam']

                    for assessment_type in assessment_types_list:
                        type_assessments = [a for a in assessments if a['type'] == assessment_type]

                        if type_assessments:
                            type_avg = sum(a['avg_score'] for a in type_assessments) / len(type_assessments)
                            type_max = max([a['avg_score'] for a in type_assessments]) if type_assessments else 0
                            type_min = min([a['avg_score'] for a in type_assessments]) if type_assessments else 0

                            # Count students passing this type
                            students_with_type = 0
                            passing_this_type = 0

                            for student in students:
                                if assessment_type in student.grades.get('by_type', {}):
                                    students_with_type += 1
                                    if student.grades['by_type'][assessment_type] >= 60:
                                        passing_this_type += 1

                            type_pass_rate = (passing_this_type / students_with_type * 100) if students_with_type > 0 else 0

                            assessment_stats[assessment_type] = {
                                'count': len(type_assessments),
                                'average': type_avg,
                                'highest': type_max,
                                'lowest': type_min,
                                'pass_rate': type_pass_rate
                            }

                except ClassSubject.DoesNotExist:
                    messages.error(request, "Selected subject not found")

        except ClassRoom.DoesNotExist:
            messages.error(request, "Selected class not found")

    context = {
        'classes': classes,
        'subjects': subjects,
        'selected_class': selected_class,
        'selected_subject_id': selected_subject_id,
        'selected_term': selected_term,
        'selected_year': selected_year,
        'students': students,
        'assessments': assessments,
        'grades': grades,
        'avg_score': avg_score,
        'assessment_count': assessment_count,
        'pass_rate': pass_rate,
        'class_avg': class_avg,
        'grade_distribution': grade_distribution,
        'assessment_stats': assessment_stats,
        'academic_terms': ['First Term', 'Second Term', 'Third Term'],
        'academic_years': ['2022-2023', '2023-2024', '2024-2025'],
    }

    return render(request, 'assignments/gradebook.html', context)

@login_required
def class_gradebook(request, class_subject_id):
    """
    Display detailed grade view for a specific class subject.
    Shows individual assignments, grade components, and student performance.
    """
    class_subject = get_object_or_404(ClassSubject, id=class_subject_id)

    # Check permissions
    if not (is_admin(request.user) or
            (is_teacher(request.user) and
             (class_subject.teacher.user == request.user or
              class_subject.classroom.class_teacher.user == request.user))):
        messages.error(request, "You don't have permission to view this gradebook")
        return redirect('assignments:gradebook')

    students = Student.objects.filter(classroom=class_subject.classroom)
    grades = Grade.objects.filter(class_subject=class_subject)

    # Process grades by student
    student_grades = {}
    for grade in grades:
        if grade.student_id not in student_grades:
            student_grades[grade.student_id] = []
        student_grades[grade.student_id].append(grade)

    # Calculate statistics for each student
    for student in students:
        student_grade_list = student_grades.get(student.id, [])
        student.grades = student_grade_list
        if student_grade_list:
            student.average = sum(g.score for g in student_grade_list) / len(student_grade_list)
            student.highest = max(g.score for g in student_grade_list)
            student.lowest = min(g.score for g in student_grade_list)
        else:
            student.average = student.highest = student.lowest = 0

    context = {
        'class_subject': class_subject,
        'students': students,
        'grades': grades,
    }

    return render(request, 'assignments/class_gradebook.html', context)

# Student specific views
@login_required
@user_passes_test(lambda u: is_student(u))
def student_grades(request):
    """
    Display a comprehensive view of the student's grades across all subjects.
    Includes filtering by subject, performance metrics, and detailed grade listing.
    """
    try:
        student = Student.objects.get(user=request.user)

        # Get filter parameter
        subject_filter = request.GET.get('subject')

        # Get enrolled subjects and grades
        enrolled_subjects = ClassSubject.objects.filter(students=student)
        grades_query = Grade.objects.filter(
            Q(submission__student=student) | Q(student=student, grade_type='TERM')
        ).select_related('class_subject', 'class_subject__subject', 'submission', 'submission__assignment')

        # Apply subject filter if specified
        if subject_filter:
            grades_query = grades_query.filter(class_subject__subject__id=subject_filter)

        # Get all grades sorted by date (newest first)
        grades = grades_query.order_by('-created_at')

        # Process grades to include percentage and ensure letter grade
        for grade in grades:
            if grade.submission and grade.submission.assignment:
                max_score = grade.submission.assignment.max_score
            else:
                max_score = grade.max_score or 100

            grade.score_percentage = (grade.score / max_score) * 100

            # Set letter grade if not already set
            if not grade.letter_grade:
                if grade.score_percentage >= 90:
                    grade.letter_grade = 'A'
                elif grade.score_percentage >= 80:
                    grade.letter_grade = 'B'
                elif grade.score_percentage >= 70:
                    grade.letter_grade = 'C'
                elif grade.score_percentage >= 60:
                    grade.letter_grade = 'D'
                else:
                    grade.letter_grade = 'F'

        # Calculate statistics
        graded_count = len(grades)

        # Pending grades (submissions without grades)
        submissions = StudentSubmission.objects.filter(
            student=student,
            is_graded=False
        ).count()
        pending_count = submissions

        # Calculate average and highest grade
        if grades:
            grade_percentages = [g.score_percentage for g in grades]
            average_grade = sum(grade_percentages) / len(grade_percentages)
            highest_grade = max(grade_percentages)
        else:
            average_grade = 0
            highest_grade = 0

        # Calculate subject performance
        subject_performance = {}
        for subject in enrolled_subjects:
            subject_grades = [g for g in grades if g.class_subject.subject.id == subject.subject.id]

            if subject_grades:
                # Calculate average for this subject
                subject_percentages = [g.score_percentage for g in subject_grades]
                subject_avg = sum(subject_percentages) / len(subject_percentages)
                subject_highest = max(subject_percentages)

                # Count grade distribution
                grade_dist = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}
                for g in subject_grades:
                    grade_dist[g.letter_grade] += 1

                # Convert to percentages
                total_grades = len(subject_grades)
                for letter in grade_dist:
                    if total_grades > 0:
                        grade_dist[letter] = (grade_dist[letter] / total_grades) * 100
                    else:
                        grade_dist[letter] = 0

                # Get completed/total assignments for this subject
                total_assignments = Assignment.objects.filter(class_subject=subject).count()
                completed_assignments = StudentSubmission.objects.filter(
                    student=student,
                    assignment__class_subject=subject
                ).count()

                # Store performance data
                subject_performance[subject.subject.name] = {
                    'average': subject_avg,
                    'highest': subject_highest,
                    'grade_distribution': grade_dist,
                    'completed': completed_assignments,
                    'total': total_assignments
                }

        context = {
            'student': student,
            'enrolled_subjects': enrolled_subjects,
            'grades': grades,
            'subject_filter': subject_filter,
            'graded_count': graded_count,
            'pending_count': pending_count,
            'average_grade': average_grade,
            'highest_grade': highest_grade,
            'subject_performance': subject_performance,
        }

        return render(request, 'assignments/student_grades.html', context)

    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('dashboard:index')

@login_required
@user_passes_test(lambda u: is_student(u))
def student_assignments(request):
    """
    Display a list of assignments for the student.
    Shows pending, submitted, and graded assignments with filtering options.
    """
    try:
        student = Student.objects.get(user=request.user)

        # Get filter parameters
        assignment_type = request.GET.get('type', 'all')
        status_filter = request.GET.get('status')

        # Get enrolled subjects
        enrolled_subjects = ClassSubject.objects.filter(students=student)

        # Get all assignments for these subjects
        assignments = Assignment.objects.filter(class_subject__in=enrolled_subjects)

        # Apply type filter if specified
        if assignment_type != 'all':
            assignments = assignments.filter(assignment_type=assignment_type)

        # Get submissions for this student
        student_submissions = StudentSubmission.objects.filter(student=student)
        submission_map = {sub.assignment_id: sub for sub in student_submissions}

        # Current time for comparing due dates
        now = timezone.now()

        # Process each assignment to add submission status
        processed_assignments = []
        for assignment in assignments:
            # Add submission if exists
            assignment.submission = submission_map.get(assignment.id)
            processed_assignments.append(assignment)

        # Apply status filter if specified
        if status_filter:
            if status_filter == 'pending':
                # No submission and not overdue
                processed_assignments = [a for a in processed_assignments if not a.submission and a.due_date >= now]
            elif status_filter == 'submitted':
                # Submitted but not graded
                processed_assignments = [a for a in processed_assignments if a.submission and not a.submission.is_graded]
            elif status_filter == 'graded':
                # Submitted and graded
                processed_assignments = [a for a in processed_assignments if a.submission and a.submission.is_graded]
            elif status_filter == 'overdue':
                # No submission and past due date
                processed_assignments = [a for a in processed_assignments if not a.submission and a.due_date < now]

        # Sort assignments by due date (closest deadline first)
        processed_assignments.sort(key=lambda a: a.due_date)

        # Calculate statistics
        total_count = len(processed_assignments)
        submitted_count = sum(1 for a in processed_assignments if a.submission)
        pending_count = sum(1 for a in processed_assignments if not a.submission and a.due_date >= now)
        overdue_count = sum(1 for a in processed_assignments if not a.submission and a.due_date < now)

        # Set section title based on filter
        section_title = "My Assignments"
        if status_filter:
            if status_filter == 'pending':
                section_title = "Pending Assignments"
            elif status_filter == 'submitted':
                section_title = "Submitted Assignments"
            elif status_filter == 'graded':
                section_title = "Graded Assignments"
            elif status_filter == 'overdue':
                section_title = "Overdue Assignments"
        elif assignment_type != 'all':
            type_mapping = {
                'HOMEWORK': 'Homework',
                'QUIZ': 'Quizzes',
                'TEST': 'Tests',
                'PROJECT': 'Projects'
            }
            section_title = type_mapping.get(assignment_type, "My Assignments")

        context = {
            'student': student,
            'assignments': processed_assignments,
            'filter_type': assignment_type,
            'status_filter': status_filter,
            'section_title': section_title,
            'total_count': total_count,
            'submitted_count': submitted_count,
            'pending_count': pending_count,
            'overdue_count': overdue_count,
            'now': now,
        }

        return render(request, 'assignments/student_assignments.html', context)

    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('dashboard:index')

# Parent specific views
@login_required
@user_passes_test(lambda u: is_parent(u))
def parent_assignments(request):
    """
    Display a list of assignments for a parent's child.
    Similar to student_assignments but for parent viewing their child's assignments.
    """
    # Get the child ID from the request
    child_id = request.GET.get('child_id')

    # Check if child_id is provided
    if not child_id:
        messages.error(request, "Please select a child to view assignments")
        return redirect('dashboard:parent_dashboard')

    try:
        # Verify the parent has access to this child
        parent = Parent.objects.get(user=request.user)
        child = get_object_or_404(Student, id=child_id)

        if not parent.children.filter(id=child.id).exists():
            messages.error(request, "You don't have permission to view this student's assignments")
            return redirect('dashboard:parent_dashboard')

        # Get filter parameters
        assignment_type = request.GET.get('type', 'all')
        status_filter = request.GET.get('status')

        # Get enrolled subjects
        enrolled_subjects = ClassSubject.objects.filter(students=child)

        # Get all assignments for these subjects
        assignments = Assignment.objects.filter(class_subject__in=enrolled_subjects)

        # Apply type filter if specified
        if assignment_type != 'all':
            assignments = assignments.filter(assignment_type=assignment_type)

        # Get submissions for this student
        student_submissions = StudentSubmission.objects.filter(student=child)
        submission_map = {sub.assignment_id: sub for sub in student_submissions}

        # Current time for comparing due dates
        now = timezone.now()

        # Process each assignment to add submission status
        processed_assignments = []
        for assignment in assignments:
            # Add submission if exists
            assignment.submission = submission_map.get(assignment.id)
            processed_assignments.append(assignment)

        # Apply status filter if specified
        if status_filter:
            if status_filter == 'pending':
                # No submission and not overdue
                processed_assignments = [a for a in processed_assignments if not a.submission and a.due_date >= now]
            elif status_filter == 'submitted':
                # Submitted but not graded
                processed_assignments = [a for a in processed_assignments if a.submission and not a.submission.is_graded]
            elif status_filter == 'graded':
                # Submitted and graded
                processed_assignments = [a for a in processed_assignments if a.submission and a.submission.is_graded]
            elif status_filter == 'overdue':
                # No submission and past due date
                processed_assignments = [a for a in processed_assignments if not a.submission and a.due_date < now]

        # Sort assignments by due date (closest deadline first)
        processed_assignments.sort(key=lambda a: a.due_date)

        # Calculate statistics
        total_count = len(processed_assignments)
        submitted_count = sum(1 for a in processed_assignments if a.submission)
        pending_count = sum(1 for a in processed_assignments if not a.submission and a.due_date >= now)
        overdue_count = sum(1 for a in processed_assignments if not a.submission and a.due_date < now)

        # Set section title based on filter
        section_title = f"{child.user.first_name}'s Assignments"
        if status_filter:
            if status_filter == 'pending':
                section_title = "Pending Assignments"
            elif status_filter == 'submitted':
                section_title = "Submitted Assignments"
            elif status_filter == 'graded':
                section_title = "Graded Assignments"
            elif status_filter == 'overdue':
                section_title = "Overdue Assignments"
        elif assignment_type != 'all':
            type_mapping = {
                'HOMEWORK': 'Homework',
                'QUIZ': 'Quizzes',
                'TEST': 'Tests',
                'PROJECT': 'Projects'
            }
            section_title = f"{type_mapping.get(assignment_type, 'Assignments')}"

        context = {
            'child': child,
            'assignments': processed_assignments,
            'filter_type': assignment_type,
            'status_filter': status_filter,
            'section_title': section_title,
            'total_count': total_count,
            'submitted_count': submitted_count,
            'pending_count': pending_count,
            'overdue_count': overdue_count,
            'now': now,
        }

        return render(request, 'assignments/parent_assignments.html', context)

    except Parent.DoesNotExist:
        messages.error(request, "Parent profile not found.")
        return redirect('dashboard:parent_dashboard')
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('dashboard:parent_dashboard')

@login_required
@user_passes_test(lambda u: is_parent(u))
def parent_grades(request):
    """
    Display a comprehensive view of a parent's child's grades across all subjects.
    Similar to student_grades but for parent viewing their child's grades.
    """
    # Get the child ID from the request
    child_id = request.GET.get('child_id')

    # Check if child_id is provided
    if not child_id:
        messages.error(request, "Please select a child to view grades")
        return redirect('dashboard:parent_dashboard')

    try:
        # Verify the parent has access to this child
        parent = Parent.objects.get(user=request.user)
        child = get_object_or_404(Student, id=child_id)

        if not parent.children.filter(id=child.id).exists():
            messages.error(request, "You don't have permission to view this student's grades")
            return redirect('dashboard:parent_dashboard')

        # Get filter parameter
        subject_filter = request.GET.get('subject')

        # Get enrolled subjects and grades
        enrolled_subjects = ClassSubject.objects.filter(students=child)
        grades_query = Grade.objects.filter(
            Q(submission__student=child) | Q(student=child)
        ).select_related('class_subject', 'class_subject__subject', 'submission', 'submission__assignment')

        # Apply subject filter if specified
        if subject_filter:
            grades_query = grades_query.filter(class_subject__subject__id=subject_filter)

        # Get all grades sorted by date (newest first)
        grades = grades_query.order_by('-created_at')

        # Process grades to include percentage and ensure letter grade
        for grade in grades:
            if grade.submission and grade.submission.assignment:
                max_score = grade.submission.assignment.max_score
            else:
                max_score = grade.max_score or 100

            grade.score_percentage = (grade.score / max_score) * 100

            # Set letter grade if not already set
            if not grade.letter_grade:
                if grade.score_percentage >= 90:
                    grade.letter_grade = 'A'
                elif grade.score_percentage >= 80:
                    grade.letter_grade = 'B'
                elif grade.score_percentage >= 70:
                    grade.letter_grade = 'C'
                elif grade.score_percentage >= 60:
                    grade.letter_grade = 'D'
                else:
                    grade.letter_grade = 'F'

        # Calculate statistics
        graded_count = len(grades)

        # Pending grades (submissions without grades)
        submissions = StudentSubmission.objects.filter(
            student=child,
            is_graded=False
        ).count()
        pending_count = submissions

        # Calculate average and highest grade
        if grades:
            grade_percentages = [g.score_percentage for g in grades]
            average_grade = sum(grade_percentages) / len(grade_percentages)
            highest_grade = max(grade_percentages)
        else:
            average_grade = 0
            highest_grade = 0

        # Calculate subject performance
        subject_performance = {}
        for subject in enrolled_subjects:
            subject_grades = [g for g in grades if g.class_subject.subject.id == subject.subject.id]

            if subject_grades:
                # Calculate average for this subject
                subject_percentages = [g.score_percentage for g in subject_grades]
                subject_avg = sum(subject_percentages) / len(subject_percentages)
                subject_highest = max(subject_percentages)

                # Count grade distribution
                grade_dist = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}
                for g in subject_grades:
                    grade_dist[g.letter_grade] += 1

                # Convert to percentages
                total_grades = len(subject_grades)
                for letter in grade_dist:
                    if total_grades > 0:
                        grade_dist[letter] = (grade_dist[letter] / total_grades) * 100
                    else:
                        grade_dist[letter] = 0

                # Get completed/total assignments for this subject
                total_assignments = Assignment.objects.filter(class_subject=subject).count()
                completed_assignments = StudentSubmission.objects.filter(
                    student=child,
                    assignment__class_subject=subject
                ).count()

                # Store performance data
                subject_performance[subject.subject.name] = {
                    'average': subject_avg,
                    'highest': subject_highest,
                    'grade_distribution': grade_dist,
                    'completed': completed_assignments,
                    'total': total_assignments
                }

        context = {
            'child': child,
            'enrolled_subjects': enrolled_subjects,
            'grades': grades,
            'subject_filter': subject_filter,
            'graded_count': graded_count,
            'pending_count': pending_count,
            'average_grade': average_grade,
            'highest_grade': highest_grade,
            'subject_performance': subject_performance,
        }

        return render(request, 'assignments/parent_grades.html', context)

    except Parent.DoesNotExist:
        messages.error(request, "Parent profile not found.")
        return redirect('dashboard:parent_dashboard')
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('dashboard:parent_dashboard')

# API views for AJAX functionality
@login_required
def api_grade_assignment(request, submission_id):
    return render(request, 'assignments/api_grade_assignment.html')

@login_required
def api_toggle_feedback(request, grade_id):
    return render(request, 'assignments/api_toggle_feedback.html')


@login_required
@user_passes_test(lambda u: is_teacher(u) or is_admin(u))
def enter_term_grades(request):
    """
    Allow teachers to enter term grades for their subjects.
    Supports dynamic assessment components based on admin-configured weight settings.
    """
    # Get the teacher's assigned subjects
    if is_teacher(request.user):
        teacher = get_object_or_404(Teacher, user=request.user)
        class_subjects = ClassSubject.objects.filter(
            Q(teacher=teacher) | Q(classroom__class_teacher=teacher)
        ).select_related('classroom', 'subject').distinct()
    else:  # Admin
        class_subjects = ClassSubject.objects.all().select_related('classroom', 'subject')

    # Get filter parameters
    selected_subject_id = request.GET.get('class_subject')
    selected_term = request.GET.get('term')
    selected_year = request.GET.get('academic_year')

    # Default to current academic year if not specified
    if not selected_year:
        from datetime import datetime
        current_year = datetime.now().year
        if datetime.now().month < 9:  # Before September
            selected_year = f"{current_year-1}-{current_year}"
        else:
            selected_year = f"{current_year}-{current_year+1}"

    students = []
    selected_subject = None
    assessment_config = None

    # Get active assessment weight configuration
    # First check for term-specific configuration, then fall back to default
    if selected_term:
        assessment_config = AssessmentWeightConfiguration.objects.filter(
            academic_term=selected_term,
            is_default=False
        ).first()

    # If no term-specific config found, use the default
    if not assessment_config:
        assessment_config = AssessmentWeightConfiguration.objects.filter(
            is_default=True
        ).first()

    # If still no config, create a default one
    if not assessment_config:
        # Create basic default configuration with traditional weights
        assessment_config = {
            'include_classwork': True,
            'classwork_weight': 10,
            'include_quizzes': False,
            'quiz_weight': 0,
            'include_tests': False,
            'test_weight': 0,
            'include_midterm': True,
            'midterm_weight': 20,
            'include_projects': True,
            'project_weight': 10,
            'include_final_exam': True,
            'final_exam_weight': 60,
            'is_default': True
        }

    # Get enabled assessment components and their weights
    enabled_components = []
    if hasattr(assessment_config, 'include_classwork') and assessment_config.include_classwork:
        enabled_components.append({
            'name': 'classwork',
            'label': 'Classwork',
            'weight': assessment_config.classwork_weight,
            'field_name': 'classwork_scores'
        })

    if hasattr(assessment_config, 'include_quizzes') and assessment_config.include_quizzes:
        enabled_components.append({
            'name': 'quiz',
            'label': 'Quizzes',
            'weight': assessment_config.quiz_weight,
            'field_name': 'quiz_scores'
        })

    if hasattr(assessment_config, 'include_tests') and assessment_config.include_tests:
        enabled_components.append({
            'name': 'test',
            'label': 'Tests',
            'weight': assessment_config.test_weight,
            'field_name': 'test_scores'
        })

    if hasattr(assessment_config, 'include_midterm') and assessment_config.include_midterm:
        enabled_components.append({
            'name': 'midterm',
            'label': 'Midterm Exams',
            'weight': assessment_config.midterm_weight,
            'field_name': 'midterm_scores'
        })

    if hasattr(assessment_config, 'include_projects') and assessment_config.include_projects:
        enabled_components.append({
            'name': 'project',
            'label': 'Project Work',
            'weight': assessment_config.project_weight,
            'field_name': 'project_scores'
        })

    if hasattr(assessment_config, 'include_final_exam') and assessment_config.include_final_exam:
        enabled_components.append({
            'name': 'exam',
            'label': 'End of Term Exams',
            'weight': assessment_config.final_exam_weight,
            'field_name': 'exam_scores'
        })

    # Ensure weights sum to 100%
    total_weight = sum(component['weight'] for component in enabled_components)
    if total_weight != 100 and total_weight > 0:
        # Normalize weights
        for component in enabled_components:
            component['weight'] = round((component['weight'] / total_weight) * 100, 1)

    if selected_subject_id:
        try:
            selected_subject = class_subjects.get(id=selected_subject_id)
            students = Student.objects.filter(enrolled_subjects=selected_subject)

            # Get existing grades for these students
            if selected_term:
                # Combine term and academic year in the term field (e.g., "First Term 2023-2024")
                term_with_year = f"{selected_term} {selected_year}"
                existing_grades = Grade.objects.filter(
                    class_subject=selected_subject,
                    student__in=students,
                    grade_type='TERM',
                    term=term_with_year
                )

                # Create a lookup dictionary for existing grades
                grade_lookup = {grade.student_id: grade for grade in existing_grades}

                # Attach existing grades to students and parse assessment scores from comments
                for student in students:
                    grade = grade_lookup.get(student.id)
                    if grade:
                        student.current_grade = grade

                        # Try to extract component scores from comments
                        if grade.comments:
                            try:
                                # Parse the JSON-like structure from comments
                                component_scores = {}

                                # Extract components using the format "Component: score, "
                                for component in enabled_components:
                                    component_label = f"{component['label']}:"
                                    if component_label in grade.comments:
                                        start_idx = grade.comments.find(component_label) + len(component_label)
                                        end_idx = grade.comments.find(",", start_idx)
                                        if end_idx == -1:  # Last component without comma
                                            end_idx = grade.comments.find("|", start_idx)
                                            if end_idx == -1:  # No trailing pipe either
                                                end_idx = len(grade.comments)

                                        score_str = grade.comments[start_idx:end_idx].strip()
                                        try:
                                            component_scores[component['name']] = float(score_str)
                                        except ValueError:
                                            # Skip invalid scores
                                            pass

                                # Extract original comment if any
                                original_comment = ""
                                if "|" in grade.comments:
                                    original_comment = grade.comments.split("|", 1)[1].strip()

                                # Add to student object for template
                                student.current_grade.component_scores = component_scores
                                student.current_grade.comments = original_comment
                            except (ValueError, IndexError):
                                # If parsing fails, initialize empty component scores
                                student.current_grade.component_scores = {}
                    else:
                        student.current_grade = None

            if request.method == 'POST':
                student_ids = request.POST.getlist('student_ids[]')
                comments = request.POST.getlist('comments[]')

                # Get scores for each assessment component
                component_scores_by_student = {}
                for component in enabled_components:
                    field_name = component['field_name']
                    scores = request.POST.getlist(f"{field_name}[]")

                    # Associate scores with students
                    for i, student_id in enumerate(student_ids):
                        if student_id not in component_scores_by_student:
                            component_scores_by_student[student_id] = {}

                        if i < len(scores) and scores[i].strip():
                            try:
                                component_scores_by_student[student_id][component['name']] = float(scores[i])
                            except ValueError:
                                # Skip invalid scores
                                pass

                # Combine term and year for the term field
                term_with_year = f"{selected_term} {selected_year}"

                # Process each student's grades
                for i, student_id in enumerate(student_ids):
                    student = get_object_or_404(Student, id=student_id)
                    comment = comments[i] if i < len(comments) else ""

                    # Skip if no scores entered for this student
                    if student_id not in component_scores_by_student or not component_scores_by_student[student_id]:
                        continue

                    # Calculate weighted total score
                    total_score = 0
                    component_comment_parts = []

                    for component in enabled_components:
                        if component['name'] in component_scores_by_student[student_id]:
                            component_score = component_scores_by_student[student_id][component['name']]
                            weighted_score = (float(component_score) * float(component['weight'])) / 100
                            total_score += weighted_score

                            # Add to comment for storage
                            component_comment_parts.append(f"{component['label']}: {component_score}")

                    # Calculate letter grade
                    if total_score >= 90:
                        letter_grade = 'A'
                    elif total_score >= 80:
                        letter_grade = 'B'
                    elif total_score >= 70:
                        letter_grade = 'C'
                    elif total_score >= 60:
                        letter_grade = 'D'
                    else:
                        letter_grade = 'F'

                    # Store component scores in comments for future reference
                    grade_comment = ", ".join(component_comment_parts)
                    if comment:
                        grade_comment += f" | {comment}"

                    # Update or create grade
                    Grade.objects.update_or_create(
                        student=student,
                        class_subject=selected_subject,
                        term=term_with_year,
                        grade_type='TERM',
                        defaults={
                            'score': total_score,
                            'max_score': 100,
                            'letter_grade': letter_grade,
                            'comments': grade_comment,
                            'created_by': request.user
                        }
                    )

                messages.success(request, f"Term grades saved successfully for {selected_subject.subject.name}")
                return redirect('assignments:enter_term_grades')

        except ClassSubject.DoesNotExist:
            messages.error(request, "Selected subject not found")
            return redirect('assignments:enter_term_grades')

    # Get academic years (current year and previous/next)
    from datetime import datetime
    current_year = datetime.now().year
    academic_years = [
        f"{year-1}-{year}" for year in range(current_year-1, current_year+2)
    ]

    context = {
        'class_subjects': class_subjects,
        'selected_subject': selected_subject,
        'selected_term': selected_term,
        'selected_year': selected_year,
        'academic_years': academic_years,
        'students': students,
        'enabled_components': enabled_components,
        'assessment_config': assessment_config,
    }

    return render(request, 'assignments/enter_term_grades.html', context)


@login_required
@user_passes_test(lambda u: is_teacher(u) or is_admin(u))
def api_auto_grade_quiz(request, submission_id):
    """
    API endpoint for auto-grading individual quiz questions via AJAX.
    Returns JSON response with the grading result.
    """
    from django.http import JsonResponse

    # Get the answer ID from the query parameters
    answer_id = request.GET.get('answer_id')

    if not answer_id:
        return JsonResponse({
            'success': False,
            'message': 'No answer ID provided'
        })

    try:
        # Get the answer and verify permissions
        answer = StudentAnswer.objects.get(id=answer_id)
        submission = answer.submission

        # Security check for teachers - verify appropriate permissions
        if is_teacher(request.user):
            try:
                teacher = Teacher.objects.get(user=request.user)

                # Check if this is a subject teacher (not a class teacher for any class)
                is_subject_teacher = not ClassRoom.objects.filter(class_teacher=teacher).exists()

                if is_subject_teacher:
                    # Subject teachers can only grade submissions for their assigned subjects
                    if not ClassSubject.objects.filter(
                        teacher=teacher,
                        id=submission.assignment.class_subject.id
                    ).exists():
                        return JsonResponse({
                            'success': False,
                            'message': 'You do not have permission to grade this submission - subject teachers can only grade their own subjects'
                        })
                else:
                    # Class teachers can grade submissions for their assigned classes
                    classroom = submission.assignment.class_subject.classroom
                    if not ClassRoom.objects.filter(class_teacher=teacher, id=classroom.id).exists():
                        return JsonResponse({
                            'success': False,
                            'message': 'You do not have permission to grade this submission - class teachers can only grade their own classes'
                        })
            except Teacher.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Teacher profile not found'
                })

        # Only auto-grade MCQ questions
        if answer.question.question_type != 'MCQ':
            return JsonResponse({
                'success': False,
                'message': 'Only multiple choice questions can be auto-graded'
            })

        # Get the correct choice
        try:
            correct_choice = Choice.objects.get(question=answer.question, is_correct=True)

            # Compare with student's choice
            if answer.selected_choice and answer.selected_choice.id == correct_choice.id:
                # Correct answer
                answer.is_correct = True
                answer.score = answer.question.points

                if answer.question.show_feedback:
                    answer.feedback = "Correct answer!"

                answer.save()

                return JsonResponse({
                    'success': True,
                    'is_correct': True,
                    'score': float(answer.score),
                    'max_score': float(answer.question.points),
                    'feedback': answer.feedback
                })
            else:
                # Incorrect answer
                answer.is_correct = False
                answer.score = 0

                if answer.question.show_feedback:
                    answer.feedback = f"Incorrect. The correct answer is: {correct_choice.choice_text}"

                answer.save()

                return JsonResponse({
                    'success': True,
                    'is_correct': False,
                    'score': 0,
                    'max_score': float(answer.question.points),
                    'feedback': answer.feedback
                })

        except Choice.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'No correct answer defined for this question'
            })

    except StudentAnswer.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Answer not found'
        })

